import unittest
import os
import sqlite3
import tempfile
from io import BytesIO

from openpyxl import load_workbook

from app import create_app
from app.routes import customers
from app.routes import reports
from app.services.report_service import (
    REPORT30_HEADERS,
    REPORT_MODE_NORMAL,
    RETROACTIVE_ENFORCEMENT_REASON_CODE,
    RETROACTIVE_JUDGMENT_REASON_CODE,
    _build_customer_as_of_report_date,
    _build_enforcement_remark,
    _build_report30_row_from_db,
    _future_effective_reason,
    apply_correction_warning_remark,
    build_correction_summary,
    build_retroactive_alerts,
    build_retroactive_enforcement_alert,
    build_retroactive_judgment_alert,
)


class ReportRetroactiveCorrectionTests(unittest.TestCase):
    def _create_report_route_db(self):
        temp_dir = tempfile.TemporaryDirectory()
        db_path = os.path.join(temp_dir.name, 'report-route.db')
        conn = sqlite3.connect(db_path)
        conn.executescript('''
            CREATE TABLE users (
                id INTEGER PRIMARY KEY,
                username TEXT,
                display_name TEXT,
                role TEXT
            );
            CREATE TABLE customers (
                id INTEGER PRIMARY KEY,
                account_no TEXT UNIQUE NOT NULL,
                name TEXT,
                case_status TEXT,
                filing_date TEXT,
                filing_capital REAL DEFAULT 0,
                judgment_date TEXT,
                total_debt REAL DEFAULT 0,
                principal REAL DEFAULT 0,
                interest_rate REAL DEFAULT 0,
                default_interest_rate REAL DEFAULT 0,
                court_fee REAL DEFAULT 0,
                lawyer_fee REAL DEFAULT 0,
                installment_count INTEGER DEFAULT 0,
                first_due_date TEXT,
                installment_1 REAL DEFAULT 0,
                installment_2 REAL DEFAULT 0,
                installment_3 REAL DEFAULT 0,
                installment_4 REAL DEFAULT 0,
                enforcement_judgment_date TEXT,
                enforcement_received_date TEXT,
                enforcement_recorded_at TEXT,
                is_deleted INTEGER DEFAULT 0
            );
            CREATE TABLE payments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                account_no TEXT,
                payment_date TEXT,
                amount REAL
            );
            CREATE TABLE case_status_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                account_no TEXT,
                from_status TEXT,
                to_status TEXT,
                changed_by INTEGER,
                changed_at TEXT,
                note TEXT
            );
            CREATE TABLE report_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                generated_by INTEGER,
                report_date TEXT,
                filename TEXT,
                status_types TEXT,
                count_30 INTEGER,
                count_31 INTEGER,
                count_33 INTEGER,
                count_alerts INTEGER,
                count_missing INTEGER,
                count_skipped INTEGER,
                count_db_total INTEGER,
                count_generated_total INTEGER,
                generated_at DATETIME DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE report_retroactive_fix_marks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                account_no TEXT NOT NULL,
                affected_report_month TEXT NOT NULL,
                effective_date TEXT NOT NULL,
                reason_code TEXT NOT NULL,
                source_report_month TEXT,
                marked_by INTEGER NOT NULL,
                marked_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                note TEXT,
                UNIQUE(account_no, affected_report_month, reason_code)
            );
        ''')
        conn.execute("INSERT INTO users VALUES (1, 'admin', 'Admin', 'admin')")

        rows = [
            (1, 'NORMAL', 'Normal', 'ยื่นฟ้อง', '2026-04-01', 100, None, 0, 0, 0, 0, 0, 0, 0, None, 0, 0, 0, 0, None, None, None, 0),
            (2, 'J-PENDING', 'Judgment Pending', 'พิพากษาตามยอม', '2026-03-10', 100, '2026-04-15', 100, 100, 0, 0, 0, 0, 2, '2026-05-30', 10, 0, 0, 0, None, None, None, 0),
            (3, 'J-MARKED', 'Judgment Marked', 'พิพากษาตามยอม', '2026-03-10', 100, '2026-04-10', 100, 100, 0, 0, 0, 0, 2, '2026-05-30', 10, 0, 0, 0, None, None, None, 0),
            (4, 'E-PENDING', 'Enforcement Pending', 'บังคับคดี', '2026-03-01', 100, '2026-04-01', 100, 100, 0, 0, 0, 0, 2, '2026-05-30', 10, 0, 0, 0, '2026-04-20', None, '2026-05-08', 0),
            (5, 'J-DEFAULT', 'Default Judgment', 'พิพากษาฝ่ายเดียว', '2026-03-10', 100, '2026-04-12', 100, 100, 0, 0, 0, 0, 0, None, 0, 0, 0, 0, None, None, None, 0),
            (6, 'E-MARKED', 'Enforcement Marked', 'บังคับคดี', '2026-03-01', 100, '2026-04-01', 100, 100, 0, 0, 0, 0, 2, '2026-05-30', 10, 0, 0, 0, '2026-04-18', None, '2026-05-08', 0),
        ]
        conn.executemany('''
            INSERT INTO customers
            (id, account_no, name, case_status, filing_date, filing_capital,
             judgment_date, total_debt, principal, interest_rate, default_interest_rate, court_fee, lawyer_fee,
             installment_count, first_due_date, installment_1, installment_2,
             installment_3, installment_4, enforcement_judgment_date,
             enforcement_received_date, enforcement_recorded_at, is_deleted)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ''', rows)
        conn.executemany('''
            INSERT INTO case_status_logs
            (account_no, from_status, to_status, changed_by, changed_at, note)
            VALUES (?, ?, ?, 1, ?, '')
        ''', [
            ('J-PENDING', 'ยื่นฟ้อง', 'พิพากษาตามยอม', '2026-05-08'),
            ('J-MARKED', 'ยื่นฟ้อง', 'พิพากษาตามยอม', '2026-05-08'),
            ('J-DEFAULT', 'ยื่นฟ้อง', 'พิพากษาฝ่ายเดียว', '2026-05-08'),
            ('E-PENDING', 'พิพากษาตามยอม', 'บังคับคดี', '2026-05-08'),
            ('E-MARKED', 'พิพากษาตามยอม', 'บังคับคดี', '2026-05-08'),
        ])
        conn.execute('''
            INSERT INTO report_retroactive_fix_marks
            (account_no, affected_report_month, effective_date, reason_code,
             source_report_month, marked_by, note)
            VALUES ('J-MARKED', '2026-04', '2026-04-10', ?, '2026-05', 1, 'done')
        ''', (RETROACTIVE_JUDGMENT_REASON_CODE,))
        conn.execute('''
            INSERT INTO report_retroactive_fix_marks
            (account_no, affected_report_month, effective_date, reason_code,
             source_report_month, marked_by, note)
            VALUES ('E-MARKED', '2026-04', '2026-04-18', ?, '2026-05', 1, 'done')
        ''', (RETROACTIVE_ENFORCEMENT_REASON_CODE,))
        conn.commit()
        conn.close()
        return temp_dir, db_path

    def test_corrected_mode_payload_is_ignored_and_generates_normal_report(self):
        temp_dir, db_path = self._create_report_route_db()
        self.addCleanup(temp_dir.cleanup)
        app = create_app()
        app.config.update(TESTING=True, DATABASE=db_path)
        original_get_current_user = reports.get_current_user
        reports.get_current_user = lambda: {'id': 1, 'role': 'admin'}
        try:
            client = app.test_client()
            normal = client.post('/api/report/generate-db', json={
                'report_date': '2026-04-30',
            }).get_json()
            corrected = client.post('/api/report/generate-db', json={
                'report_date': '2026-04-30',
                'report_mode': 'corrected',
                'corrected_scope': 'pending_only',
            }).get_json()
            source_month = client.post('/api/report/generate-db', json={
                'report_date': '2026-05-31',
            }).get_json()

            normal_accounts = {
                row.get('account_no') for row in normal['report_30'] + normal['report_31']
            }
            corrected_accounts = {
                row.get('account_no') for row in corrected['report_30'] + corrected['report_31']
            }

            self.assertIn('NORMAL', normal_accounts)
            self.assertIn('J-MARKED', normal_accounts)
            self.assertEqual(corrected['report_mode'], 'normal')
            self.assertEqual(corrected_accounts, normal_accounts)
            self.assertIn('J-PENDING', corrected_accounts)
            self.assertIn('E-PENDING', corrected_accounts)

            j_pending = next(row for row in corrected['report_31'] if row.get('account_no') == 'J-PENDING')
            e_pending = next(row for row in corrected['report_30'] if row.get('account_no') == 'E-PENDING')
            j_marked = next(row for row in corrected['report_31'] if row.get('account_no') == 'J-MARKED')
            j_default = next(row for row in corrected['report_30'] if row.get('account_no') == 'J-DEFAULT')

            self.assertNotIn('ย้อนหลัง', j_pending['remark'])
            self.assertEqual(e_pending['case_status'], 'บังคับคดี')
            self.assertNotIn('มีหมายบังคับคดีย้อนหลัง', e_pending['report30_litigation_remark'])
            self.assertNotIn('ย้อนหลัง', j_marked.get('remark') or '')
            self.assertEqual(j_default['report30_note_2'], 'พิพากษาฝ่ายเดียว')
            self.assertNotIn('ย้อนหลัง', j_default.get('report30_litigation_remark') or '')
            source_j_pending = next(row for row in source_month['report_31'] if row.get('account_no') == 'J-PENDING')
            source_e_pending = next(row for row in source_month['report_30'] if row.get('account_no') == 'E-PENDING')
            source_e_marked = next(row for row in source_month['report_30'] if row.get('account_no') == 'E-MARKED')
            source_j_marked = next(row for row in source_month['report_31'] if row.get('account_no') == 'J-MARKED')
            self.assertIn('มีคำพิพากษาย้อนหลัง', source_j_pending['remark'])
            self.assertIn('04/2026', source_j_pending['remark'])
            self.assertIn('เป็นต้นไป', source_j_pending['remark'])
            self.assertIn('มีหมายบังคับคดีย้อนหลัง', source_e_pending['report30_litigation_remark'])
            self.assertIn('04/2026', source_e_pending['report30_litigation_remark'])
            self.assertIn('เป็นต้นไป', source_e_pending['report30_litigation_remark'])
            self.assertNotIn('ย้อนหลัง', source_e_marked.get('report30_litigation_remark') or '')
            self.assertNotIn('ย้อนหลัง', source_j_marked.get('remark') or '')
        finally:
            reports.get_current_user = original_get_current_user

    def test_invalid_corrected_scope_is_ignored(self):
        temp_dir, db_path = self._create_report_route_db()
        self.addCleanup(temp_dir.cleanup)
        app = create_app()
        app.config.update(TESTING=True, DATABASE=db_path)
        original_get_current_user = reports.get_current_user
        reports.get_current_user = lambda: {'id': 1, 'role': 'admin'}
        try:
            client = app.test_client()
            response = client.post('/api/report/generate-db', json={
                'report_date': '2026-04-30',
                'report_mode': 'corrected',
                'corrected_scope': 'all_accounts',
            })

            self.assertEqual(response.status_code, 200)
            self.assertEqual(response.get_json()['report_mode'], 'normal')
        finally:
            reports.get_current_user = original_get_current_user

    def test_corrected_mode_payload_does_not_filter_to_pending_only(self):
        temp_dir, db_path = self._create_report_route_db()
        self.addCleanup(temp_dir.cleanup)
        app = create_app()
        app.config.update(TESTING=True, DATABASE=db_path)
        original_get_current_user = reports.get_current_user
        reports.get_current_user = lambda: {'id': 1, 'role': 'admin'}
        try:
            client = app.test_client()
            response = client.post('/api/report/generate-db', json={
                'report_date': '2026-06-30',
                'report_mode': 'corrected',
                'corrected_scope': 'pending_only',
            })
            payload = response.get_json()

            self.assertEqual(response.status_code, 200)
            self.assertEqual(payload['report_mode'], 'normal')
            self.assertGreater(payload['summary']['generated_total'], 0)
        finally:
            reports.get_current_user = original_get_current_user

    def test_marked_corrections_are_excluded_from_pending_total(self):
        temp_dir, db_path = self._create_report_route_db()
        self.addCleanup(temp_dir.cleanup)
        app = create_app()
        app.config.update(TESTING=True, DATABASE=db_path)
        original_get_current_user = reports.get_current_user
        reports.get_current_user = lambda: {'id': 1, 'role': 'admin'}
        try:
            client = app.test_client()
            payload = client.post('/api/report/generate-db', json={
                'report_date': '2026-04-30',
            }).get_json()

            self.assertEqual(payload['correction_summary']['total'], 5)
            self.assertEqual(payload['correction_summary']['pending_total'], 3)
        finally:
            reports.get_current_user = original_get_current_user

    def test_missing_judgment_recorded_date_does_not_create_retroactive_alert(self):
        alert = build_retroactive_judgment_alert({
            'account_no': 'NO-LOG',
            'case_status': 'พิพากษาฝ่ายเดียว',
            'filing_date': '2026-03-10',
            'judgment_date': '2026-04-15',
            '_case_status_logs': [],
        })

        self.assertIsNone(alert)

    def test_manual_judgment_date_edit_to_prior_month_creates_alert(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = os.path.join(temp_dir, 'manual-judgment-edit.db')
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            conn.executescript('''
                CREATE TABLE customer_edits (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    customer_id INTEGER,
                    account_no TEXT,
                    edited_by INTEGER,
                    edited_at TEXT,
                    changes TEXT
                );
                CREATE TABLE report_retroactive_fix_marks (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    account_no TEXT NOT NULL,
                    affected_report_month TEXT NOT NULL,
                    effective_date TEXT NOT NULL,
                    reason_code TEXT NOT NULL,
                    source_report_month TEXT,
                    marked_by INTEGER NOT NULL,
                    marked_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    note TEXT,
                    UNIQUE(account_no, affected_report_month, reason_code)
                );
                CREATE TABLE users (
                    id INTEGER PRIMARY KEY,
                    username TEXT,
                    display_name TEXT,
                    role TEXT
                );
            ''')
            conn.execute("INSERT INTO users VALUES (1, 'admin', 'Admin', 'admin')")
            conn.execute(
                '''
                INSERT INTO customer_edits
                (customer_id, account_no, edited_by, edited_at, changes)
                VALUES (?, ?, ?, ?, ?)
                ''',
                (
                    1,
                    'J-MANUAL',
                    1,
                    '2026-05-10 09:00:00',
                    '{"judgment_date":{"label":"วันพิพากษา","from":"2026-05-15","to":"2026-04-20"}}',
                ),
            )

            alert = build_retroactive_judgment_alert({
                'account_no': 'J-MANUAL',
                'case_status': 'พิพากษาตามยอม',
                'filing_date': '2026-03-10',
                'judgment_date': '2026-04-20',
                '_case_status_logs': [],
            }, db=conn)

            self.assertIsNotNone(alert)
            self.assertEqual(alert['reason_code'], RETROACTIVE_JUDGMENT_REASON_CODE)
            self.assertEqual(alert['affected_report_month'], '2026-04')
            self.assertEqual(alert['source_report_month'], '2026-05')
            self.assertEqual(alert['recorded_date'], '2026-05-10')
            conn.close()

    def test_manual_judgment_date_edit_same_month_does_not_create_alert(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = os.path.join(temp_dir, 'same-month-judgment-edit.db')
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            conn.executescript('''
                CREATE TABLE customer_edits (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    customer_id INTEGER,
                    account_no TEXT,
                    edited_by INTEGER,
                    edited_at TEXT,
                    changes TEXT
                );
            ''')
            conn.execute(
                '''
                INSERT INTO customer_edits
                (customer_id, account_no, edited_by, edited_at, changes)
                VALUES (?, ?, ?, ?, ?)
                ''',
                (
                    1,
                    'J-SAME',
                    1,
                    '2026-05-10 09:00:00',
                    '{"judgment_date":{"label":"วันพิพากษา","from":"2026-05-15","to":"2026-05-01"}}',
                ),
            )

            alert = build_retroactive_judgment_alert({
                'account_no': 'J-SAME',
                'case_status': 'พิพากษาตามยอม',
                'filing_date': '2026-03-10',
                'judgment_date': '2026-05-01',
                '_case_status_logs': [],
            }, db=conn)

            self.assertIsNone(alert)
            conn.close()

    def test_judgment_after_report_rolls_back_to_filing(self):
        customer = {
            'case_status': 'พิพากษาฝ่ายเดียว',
            'filing_date': '2026-02-10',
            'judgment_date': '2026-04-15',
        }

        customer_as_of = _build_customer_as_of_report_date(customer, '2026-03-31')

        self.assertEqual(customer_as_of['case_status'], 'ยื่นฟ้อง')
        self.assertFalse(_future_effective_reason(customer_as_of, '2026-03-31'))

    def test_filing_after_report_still_blocks_as_filing_after_report(self):
        reason = _future_effective_reason(
            {
                'case_status': 'พิพากษาฝ่ายเดียว',
                'filing_date': '2026-04-01',
                'judgment_date': '2026-04-15',
            },
            '2026-03-31',
        )

        self.assertEqual(reason[0], 'FILING_DATE_AFTER_REPORT')

    def test_retroactive_default_judgment_creates_alert_and_source_month_remark_only(self):
        customer = {
            'account_no': 'J-1',
            'case_status': 'พิพากษาฝ่ายเดียว',
            'filing_date': '2026-03-10',
            'judgment_date': '2026-04-15',
            'total_debt': 100,
            '_case_status_logs': [
                {
                    'from_status': 'ยื่นฟ้อง',
                    'to_status': 'พิพากษาฝ่ายเดียว',
                    'changed_at': '2026-05-08',
                },
            ],
        }
        alert = build_retroactive_judgment_alert(customer)

        self.assertIsNotNone(alert)
        self.assertEqual(alert['reason_code'], RETROACTIVE_JUDGMENT_REASON_CODE)
        self.assertEqual(alert['type'], 'judgment')
        self.assertEqual(alert['from_status'], 'ยื่นฟ้อง')
        self.assertEqual(alert['to_status'], 'พิพากษาฝ่ายเดียว')
        self.assertEqual(alert['affected_report_month'], '2026-04')
        self.assertEqual(alert['source_report_month'], '2026-05')
        self.assertEqual(alert['recorded_date'], '2026-05-08')

        affected_customer_as_of = _build_customer_as_of_report_date(
            customer,
            '2026-04-30',
            report_mode=REPORT_MODE_NORMAL,
            retroactive_alerts=[alert],
        )
        affected_row = _build_report30_row_from_db(
            'J-1',
            affected_customer_as_of['case_status'],
            affected_customer_as_of['filing_date'],
            100,
            affected_customer_as_of,
            None,
            '2026-04-30',
            payments=[],
        )
        affected_row = apply_correction_warning_remark(affected_row, affected_customer_as_of)

        source_customer_as_of = _build_customer_as_of_report_date(
            customer,
            '2026-05-31',
            report_mode=REPORT_MODE_NORMAL,
            retroactive_alerts=[alert],
        )
        source_row = _build_report30_row_from_db(
            'J-1',
            source_customer_as_of['case_status'],
            source_customer_as_of['filing_date'],
            100,
            source_customer_as_of,
            None,
            '2026-05-31',
            payments=[],
        )
        source_row = apply_correction_warning_remark(source_row, source_customer_as_of)

        self.assertEqual(affected_customer_as_of['case_status'], 'พิพากษาฝ่ายเดียว')
        self.assertEqual(affected_row['report30_note_2'], 'พิพากษาฝ่ายเดียว')
        self.assertNotIn('ย้อนหลัง', affected_row['report30_litigation_remark'])
        self.assertIn('มีคำพิพากษาย้อนหลัง', source_row['report30_litigation_remark'])
        self.assertIn('04/2026', source_row['report30_litigation_remark'])
        self.assertIn('เป็นต้นไป', source_row['report30_litigation_remark'])

    def test_retroactive_consent_judgment_shows_remark_in_source_month_only(self):
        customer = {
            'account_no': 'J-31',
            'case_status': 'พิพากษาตามยอม',
            'filing_date': '2026-03-10',
            'judgment_date': '2026-04-15',
            'total_debt': 100,
            '_case_status_logs': [
                {
                    'from_status': 'ยื่นฟ้อง',
                    'to_status': 'พิพากษาตามยอม',
                    'changed_at': '2026-05-08',
                },
            ],
        }
        alert = build_retroactive_judgment_alert(customer)

        affected_customer_as_of = _build_customer_as_of_report_date(
            customer,
            '2026-04-30',
            report_mode=REPORT_MODE_NORMAL,
            retroactive_alerts=[alert],
        )
        affected_row = apply_correction_warning_remark({'remark': ''}, affected_customer_as_of)

        source_customer_as_of = _build_customer_as_of_report_date(
            customer,
            '2026-05-31',
            report_mode=REPORT_MODE_NORMAL,
            retroactive_alerts=[alert],
        )
        source_row = apply_correction_warning_remark({'remark': ''}, source_customer_as_of)

        self.assertEqual(alert['reason_code'], RETROACTIVE_JUDGMENT_REASON_CODE)
        self.assertEqual(alert['from_status'], 'ยื่นฟ้อง')
        self.assertEqual(alert['to_status'], 'พิพากษาตามยอม')
        self.assertEqual(alert['affected_report_month'], '2026-04')
        self.assertEqual(alert['source_report_month'], '2026-05')
        self.assertEqual(affected_customer_as_of['case_status'], 'พิพากษาตามยอม')
        self.assertEqual(source_customer_as_of['case_status'], 'พิพากษาตามยอม')
        self.assertNotIn('ย้อนหลัง', affected_row['remark'])
        self.assertIn('มีคำพิพากษาย้อนหลัง', source_row['remark'])
        self.assertIn('04/2026', source_row['remark'])
        self.assertIn('เป็นต้นไป', source_row['remark'])

    def test_retroactive_judgment_marked_alert_does_not_add_pending_remark(self):
        customer = {
            'account_no': 'J-2',
            'case_status': 'พิพากษาตามยอม',
            'filing_date': '2026-03-10',
            'judgment_date': '2026-04-15',
            '_case_status_logs': [
                {
                    'from_status': 'ยื่นฟ้อง',
                    'to_status': 'พิพากษาตามยอม',
                    'changed_at': '2026-05-08',
                },
            ],
        }
        alert = build_retroactive_judgment_alert(customer)
        alert['marked'] = True

        customer_as_of = _build_customer_as_of_report_date(
            customer,
            '2026-05-31',
            retroactive_alerts=[alert],
        )
        row = apply_correction_warning_remark({'remark': ''}, customer_as_of)

        self.assertEqual(customer_as_of['case_status'], 'พิพากษาตามยอม')
        self.assertNotIn('_pending_correction_alert', customer_as_of)
        self.assertEqual(row['remark'], '')

    def test_retroactive_marked_normal_mode_uses_effective_status(self):
        customer = {
            'case_status': 'พิพากษาฝ่ายเดียว',
            'filing_date': '2026-03-10',
            'judgment_date': '2026-04-15',
        }
        alert = {
            'type': 'judgment',
            'effective_date': '2026-04-15',
            'marked': True,
        }

        customer_as_of = _build_customer_as_of_report_date(
            customer,
            '2026-04-30',
            report_mode=REPORT_MODE_NORMAL,
            retroactive_alerts=[alert],
        )

        self.assertIs(customer_as_of, customer)
        self.assertEqual(customer_as_of['case_status'], 'พิพากษาฝ่ายเดียว')

    def test_retroactive_enforcement_shows_remark_in_source_month_only(self):
        customer = {
            'account_no': 'E-1',
            'case_status': 'บังคับคดี',
            'filing_date': '2026-03-01',
            'judgment_date': '2026-04-01',
            'enforcement_judgment_date': '2026-04-20',
            'enforcement_recorded_at': '2026-05-08',
            '_case_status_logs': [
                {
                    'from_status': 'พิพากษาตามยอม',
                    'to_status': 'บังคับคดี',
                    'changed_at': '2026-05-08',
                },
            ],
        }
        alert = build_retroactive_enforcement_alert(customer)

        affected_customer = _build_customer_as_of_report_date(
            customer,
            '2026-04-30',
            report_mode=REPORT_MODE_NORMAL,
            retroactive_alerts=[alert],
        )
        affected_row = apply_correction_warning_remark({'remark': 'เดิม'}, affected_customer)

        source_customer = _build_customer_as_of_report_date(
            customer,
            '2026-05-31',
            report_mode=REPORT_MODE_NORMAL,
            retroactive_alerts=[alert],
        )
        source_row = apply_correction_warning_remark({'remark': 'เดิม'}, source_customer)

        self.assertEqual(alert['reason_code'], RETROACTIVE_ENFORCEMENT_REASON_CODE)
        self.assertEqual(alert['affected_report_month'], '2026-04')
        self.assertEqual(alert['source_report_month'], '2026-05')
        self.assertEqual(affected_customer['case_status'], 'บังคับคดี')
        self.assertEqual(source_customer['case_status'], 'บังคับคดี')
        self.assertNotIn('ย้อนหลัง', affected_row['remark'])
        self.assertIn('เดิม | มีหมายบังคับคดีย้อนหลัง', source_row['remark'])
        self.assertIn('04/2026', source_row['remark'])
        self.assertIn('เป็นต้นไป', source_row['remark'])

    def test_enforcement_remark_same_month_order_and_recorded_date_is_empty(self):
        customer = {
            'account_no': 'E-SAME-MONTH',
            'case_status': 'บังคับคดี',
            'filing_date': '2026-03-01',
            'judgment_date': '2026-04-01',
            'enforcement_judgment_date': '2026-05-06',
            'enforcement_recorded_at': '2026-05-27',
            '_case_status_logs': [
                {
                    'from_status': 'พิพากษาตามยอม',
                    'to_status': 'บังคับคดี',
                    'changed_at': '2026-05-27',
                },
            ],
            '_has_consent_to_enforcement_log': True,
        }

        remark = _build_enforcement_remark(customer, {}, '2026-05-31')
        row = _build_report30_row_from_db(
            customer['account_no'],
            customer['case_status'],
            customer['filing_date'],
            100,
            customer,
            {'principal_bal': 100, 'remaining_debt_raw': 100},
            '2026-05-31',
            payments=[],
        )

        self.assertEqual(remark, '')
        self.assertEqual(row['report30_litigation_remark'], '')

    def test_enforcement_remark_uses_enforcement_dates_only_in_recorded_month(self):
        customer = {
            'account_no': 'E-RETRO',
            'case_status': 'บังคับคดี',
            'filing_date': '2026-03-01',
            'judgment_date': '2026-05-06',
            'enforcement_judgment_date': '2026-04-25',
            '_case_status_logs': [
                {
                    'from_status': 'พิพากษาตามยอม',
                    'to_status': 'บังคับคดี',
                    'changed_at': '2026-05-27',
                },
            ],
            '_has_consent_to_enforcement_log': True,
        }

        source_remark = _build_enforcement_remark(customer, {}, '2026-05-31')
        affected_remark = _build_enforcement_remark(customer, {}, '2026-04-30')

        self.assertIn('2026-04-25', source_remark)
        self.assertIn('ตั้งแต่เดือน 2026-04', source_remark)
        self.assertEqual(affected_remark, '')

    def test_enforcement_remark_marked_alert_is_empty_after_fix(self):
        customer = {
            'account_no': 'E-MARKED',
            'case_status': 'บังคับคดี',
            'filing_date': '2026-03-01',
            'judgment_date': '2026-04-01',
            'enforcement_judgment_date': '2026-04-25',
            'enforcement_recorded_at': '2026-05-27',
            '_case_status_logs': [
                {
                    'from_status': 'พิพากษาตามยอม',
                    'to_status': 'บังคับคดี',
                    'changed_at': '2026-05-27',
                },
            ],
            '_has_consent_to_enforcement_log': True,
        }
        marked_alert = {
            'type': 'enforcement',
            'marked': True,
            'source_report_month': '2026-05',
            'affected_report_month': '2026-04',
        }

        customer_as_of = _build_customer_as_of_report_date(
            customer,
            '2026-05-31',
            report_mode=REPORT_MODE_NORMAL,
            retroactive_alerts=[marked_alert],
        )

        self.assertEqual(_build_enforcement_remark(customer_as_of, {}, '2026-05-31'), '')

    def test_enforcement_remark_does_not_fallback_to_judgment_or_filing_dates(self):
        customer = {
            'account_no': 'E-NO-ENFORCEMENT-DATE',
            'case_status': 'บังคับคดี',
            'filing_date': '2026-04-01',
            'judgment_date': '2026-04-25',
            'enforcement_recorded_at': '2026-05-27',
            '_case_status_logs': [
                {
                    'from_status': 'พิพากษาตามยอม',
                    'to_status': 'บังคับคดี',
                    'changed_at': '2026-05-27',
                },
            ],
            '_has_consent_to_enforcement_log': True,
        }

        self.assertEqual(_build_enforcement_remark(customer, {}, '2026-05-31'), '')

    def test_retroactive_enforcement_from_default_judgment_does_not_create_alert(self):
        customer = {
            'account_no': 'E-DEFAULT',
            'case_status': 'บังคับคดี',
            'filing_date': '2026-02-01',
            'judgment_date': '2026-03-01',
            'enforcement_judgment_date': '2026-03-20',
            'enforcement_recorded_at': '2026-04-08',
            '_case_status_logs': [
                {
                    'from_status': 'พิพากษาฝ่ายเดียว',
                    'to_status': 'บังคับคดี',
                    'changed_at': '2026-04-08',
                },
            ],
        }

        alert = build_retroactive_enforcement_alert(customer)
        alerts = build_retroactive_alerts(customer)
        summary = build_correction_summary(alerts, '2026-03-31')

        self.assertIsNone(alert)
        self.assertFalse(any(item.get('type') == 'enforcement' for item in alerts))
        self.assertEqual(summary['pending_enforcement'], 0)

    def test_corrected_mode_payload_keeps_normal_generation_without_default_enforcement_alert(self):
        temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        db_path = os.path.join(temp_dir.name, 'default-enforcement-route.db')
        conn = sqlite3.connect(db_path)
        conn.executescript('''
            CREATE TABLE users (
                id INTEGER PRIMARY KEY,
                username TEXT,
                display_name TEXT,
                role TEXT
            );
            CREATE TABLE customers (
                id INTEGER PRIMARY KEY,
                account_no TEXT UNIQUE NOT NULL,
                name TEXT,
                case_status TEXT,
                filing_date TEXT,
                filing_capital REAL DEFAULT 0,
                judgment_date TEXT,
                total_debt REAL DEFAULT 0,
                principal REAL DEFAULT 0,
                interest_rate REAL DEFAULT 0,
                default_interest_rate REAL DEFAULT 0,
                court_fee REAL DEFAULT 0,
                lawyer_fee REAL DEFAULT 0,
                installment_count INTEGER DEFAULT 0,
                first_due_date TEXT,
                installment_1 REAL DEFAULT 0,
                installment_2 REAL DEFAULT 0,
                installment_3 REAL DEFAULT 0,
                installment_4 REAL DEFAULT 0,
                enforcement_judgment_date TEXT,
                enforcement_received_date TEXT,
                enforcement_recorded_at TEXT,
                is_deleted INTEGER DEFAULT 0
            );
            CREATE TABLE payments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                account_no TEXT,
                payment_date TEXT,
                amount REAL
            );
            CREATE TABLE case_status_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                account_no TEXT,
                from_status TEXT,
                to_status TEXT,
                changed_by INTEGER,
                changed_at TEXT,
                note TEXT
            );
            CREATE TABLE report_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                generated_by INTEGER,
                report_date TEXT,
                filename TEXT,
                status_types TEXT,
                count_30 INTEGER,
                count_31 INTEGER,
                count_33 INTEGER,
                count_alerts INTEGER,
                count_missing INTEGER,
                count_skipped INTEGER,
                count_db_total INTEGER,
                count_generated_total INTEGER,
                generated_at DATETIME DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE report_retroactive_fix_marks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                account_no TEXT NOT NULL,
                affected_report_month TEXT NOT NULL,
                effective_date TEXT NOT NULL,
                reason_code TEXT NOT NULL,
                source_report_month TEXT,
                marked_by INTEGER NOT NULL,
                marked_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                note TEXT,
                UNIQUE(account_no, affected_report_month, reason_code)
            );
        ''')
        conn.execute("INSERT INTO users VALUES (1, 'admin', 'Admin', 'admin')")
        conn.execute('''
            INSERT INTO customers
            (id, account_no, name, case_status, filing_date, filing_capital,
             judgment_date, total_debt, principal, interest_rate, default_interest_rate,
             court_fee, lawyer_fee, installment_count, first_due_date, installment_1,
             installment_2, installment_3, installment_4, enforcement_judgment_date,
             enforcement_received_date, enforcement_recorded_at, is_deleted)
            VALUES (1, 'E-DEFAULT', 'Default Enforcement', 'บังคับคดี',
                    '2026-02-01', 100, '2026-03-01', 100, 100, 0, 0,
                    0, 0, 1, '2026-03-31', 100, 0, 0, 0,
                    '2026-03-20', NULL, '2026-04-08', 0)
        ''')
        conn.execute('''
            INSERT INTO case_status_logs
            (account_no, from_status, to_status, changed_by, changed_at, note)
            VALUES ('E-DEFAULT', 'พิพากษาฝ่ายเดียว', 'บังคับคดี', 1, '2026-04-08', '')
        ''')
        conn.commit()
        conn.close()

        app = create_app()
        app.config.update(TESTING=True, DATABASE=db_path)
        original_get_current_user = reports.get_current_user
        reports.get_current_user = lambda: {'id': 1, 'role': 'admin'}
        try:
            client = app.test_client()
            corrected = client.post('/api/report/generate-db', json={
                'report_date': '2026-03-31',
                'report_mode': 'corrected',
                'corrected_scope': 'pending_only',
            }).get_json()

            generated_accounts = {
                row.get('account_no') for row in corrected['report_30'] + corrected['report_31']
            }
            self.assertIn('E-DEFAULT', generated_accounts)
            self.assertEqual(corrected['summary']['correction_summary']['pending_enforcement'], 0)
            self.assertEqual(corrected['report_mode'], 'normal')
        finally:
            reports.get_current_user = original_get_current_user

    def test_export_all_does_not_add_retroactive_alert_sheet(self):
        app = create_app()
        app.config.update(TESTING=True)
        original_get_current_user = reports.get_current_user
        reports.get_current_user = lambda: {'id': 1, 'role': 'admin'}
        try:
            with app.test_request_context('/api/report/export-all', method='POST', json={
                'report_date': '2026-04-30',
                'report_30': [{
                    'account_no': 'J-1',
                    'report30_filing_date': '20260401',
                    'report30_filing_amount': 100,
                    'report30_note': 'คดีแพ่ง',
                    'report30_note_2': 'ฟ้องแล้ว',
                    'report30_note_3': '1.1',
                    'report30_litigation_remark': 'มีคำพิพากษาย้อนหลัง',
                }],
                'retroactive_alerts': [{
                    'type': 'judgment',
                    'account_no': 'J-1',
                    'reason_code': RETROACTIVE_JUDGMENT_REASON_CODE,
                }],
            }):
                response = reports.export_all_reports()

            response.direct_passthrough = False
            workbook = load_workbook(BytesIO(response.get_data()), read_only=True)
            self.assertNotIn('Retroactive Alerts', workbook.sheetnames)
            self.assertIn('Report 30', workbook.sheetnames)
            self.assertEqual(
                [cell.value for cell in next(workbook['Report 30'].iter_rows(min_row=1, max_row=1))],
                REPORT30_HEADERS,
            )
        finally:
            reports.get_current_user = original_get_current_user

    def test_generic_retroactive_report_fix_endpoint_marks_judgment(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            db_path = os.path.join(temp_dir, 'retro-mark.db')
            conn = sqlite3.connect(db_path)
            conn.executescript('''
                CREATE TABLE users (
                    id INTEGER PRIMARY KEY,
                    username TEXT,
                    display_name TEXT,
                    role TEXT
                );
                CREATE TABLE customers (
                    id INTEGER PRIMARY KEY,
                    account_no TEXT UNIQUE NOT NULL,
                    name TEXT,
                    case_status TEXT,
                    filing_date TEXT,
                    judgment_date TEXT,
                    is_deleted INTEGER DEFAULT 0
                );
                CREATE TABLE case_status_logs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    account_no TEXT,
                    from_status TEXT,
                    to_status TEXT,
                    changed_by INTEGER,
                    changed_at TEXT,
                    note TEXT
                );
                CREATE TABLE report_retroactive_fix_marks (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    account_no TEXT NOT NULL,
                    affected_report_month TEXT NOT NULL,
                    effective_date TEXT NOT NULL,
                    reason_code TEXT NOT NULL,
                    source_report_month TEXT,
                    marked_by INTEGER NOT NULL,
                    marked_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    note TEXT,
                    UNIQUE(account_no, affected_report_month, reason_code)
                );
                CREATE TABLE customer_edits (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    customer_id INTEGER,
                    account_no TEXT,
                    edited_by INTEGER,
                    edited_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    changes TEXT
                );
            ''')
            conn.execute("INSERT INTO users VALUES (1, 'admin', 'Admin', 'admin')")
            conn.execute("""
                INSERT INTO customers
                (id, account_no, name, case_status, filing_date, judgment_date, is_deleted)
                VALUES (1, 'J-MARK', 'Judgment Mark', 'พิพากษาตามยอม', '2026-03-10', '2026-04-15', 0)
            """)
            conn.execute("""
                INSERT INTO case_status_logs
                (account_no, from_status, to_status, changed_by, changed_at, note)
                VALUES ('J-MARK', 'ยื่นฟ้อง', 'พิพากษาตามยอม', 1, '2026-05-08', '')
            """)
            conn.commit()
            conn.close()

            app = create_app()
            app.config.update(TESTING=True, DATABASE=db_path)
            original_get_current_user = customers.get_current_user
            customers.get_current_user = lambda: {'id': 1, 'role': 'admin'}
            try:
                client = app.test_client()
                response = client.post('/api/customers/J-MARK/retroactive-report-fix', json={
                    'reason_code': RETROACTIVE_JUDGMENT_REASON_CODE,
                    'affected_report_month': '2026-04',
                    'note': 'แก้รายงานย้อนหลังแล้ว',
                })
                payload = response.get_json()

                self.assertEqual(response.status_code, 200)
                self.assertTrue(payload['retroactive_alert']['marked'])
                self.assertEqual(payload['retroactive_alert']['reason_code'], RETROACTIVE_JUDGMENT_REASON_CODE)
            finally:
                customers.get_current_user = original_get_current_user


if __name__ == '__main__':
    unittest.main()
