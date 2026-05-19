import io
import json
import re
import openpyxl
from datetime import date, datetime
from flask import Blueprint, request, jsonify, send_file, current_app
from app.database import get_db
from app.services.auth_service import get_user_by_token
from app.services.status_service import refresh_customer_status
from app.services.report_service import get_snapshot_at_date
from app.services.customer_list_cache_service import refresh_customer_list_cache
from dateutil.relativedelta import relativedelta

bp = Blueprint('imports', __name__, url_prefix='/api/import')


def get_current_user():
    token = request.cookies.get(current_app.config.get('AUTH_COOKIE_NAME', 'token')) or request.headers.get('Authorization', '').replace('Bearer ', '')
    return get_user_by_token(token)


@bp.before_request
def block_superadmin_from_import_api():
    user = get_current_user()
    if user and user['role'] == 'superadmin':
        return jsonify({'error': 'Superadmin is limited to user management.'}), 403


def parse_account_no(val):
    if val is None:
        return None

    s = str(val).strip()

    if not s or s == 'None':
        return None

    if s.endswith('.0') and s[:-2].isdigit():
        s = s[:-2]

    s = s.replace('-', '').replace(' ', '')

    return s


def is_valid_customer_name(value):
    text = str(value or '').strip()
    if not text:
        return False
    if re.search(r'\s{2,}', text):
        return False
    return bool(re.fullmatch(r'[A-Za-z0-9ก-ฮะ-์.\-\s]+', text))


def parse_date(val):
    """
    Import date รับเฉพาะ Text รูปแบบ YYYY-MM-DD เท่านั้น เช่น 2026-03-12
    ไม่รับ Excel Date Object และไม่เดา DD/MM/YYYY, DD-MM-YYYY, MM/DD/YYYY
    """
    if val is None or str(val).strip() == '':
        return None

    if isinstance(val, (datetime, date)):
        return None

    s = str(val).strip()

    if not re.fullmatch(r'\d{4}-\d{2}-\d{2}', s):
        return None

    try:
        return datetime.strptime(s, '%Y-%m-%d').strftime('%Y-%m-%d')
    except Exception:
        return None


def validate_judgment_timeline(filing_date, judgment_type, judgment_date, first_due_date):
    if filing_date and judgment_date and judgment_date <= filing_date:
        raise ValueError(
            'วันที่พิพากษาต้องมากกว่าวันที่ยื่นฟ้อง '
            f'(วันที่ยื่นฟ้องในระบบ: {filing_date}, วันที่พิพากษาในไฟล์: {judgment_date})'
        )

    if judgment_date and first_due_date:
        if judgment_type == 'พิพากษาฝ่ายเดียว':
            if first_due_date < judgment_date:
                raise ValueError(
                    'วันครบกำหนดงวดแรกต้องไม่น้อยกว่าวันที่พิพากษา '
                    f'(วันที่พิพากษา: {judgment_date}, วันครบกำหนดงวดแรก: {first_due_date})'
                )
        else:
            if first_due_date <= judgment_date:
                raise ValueError(
                    'วันครบกำหนดงวดแรกต้องมากกว่าวันที่พิพากษา '
                    f'(วันที่พิพากษา: {judgment_date}, วันครบกำหนดงวดแรก: {first_due_date})'
                )


def normalize_black_case_no(value):
    raw = str(value or '').strip()
    if not raw:
        return ''

    raw = re.sub(r'\s*/\s*', '/', raw)
    raw = re.sub(r'\s+', ' ', raw)
    match = re.fullmatch(r'([A-Za-zก-ฮ]{1,8})(?:\s+([A-Za-z]?\d{1,8})|(\d{1,8}))/(25\d{2})', raw)
    if not match:
        return None

    case_type = match.group(1)
    case_no = match.group(2) or match.group(3)
    year = match.group(4)
    return f'{case_type} {case_no}/{year}'


def normalize_red_case_no(value):
    return normalize_black_case_no(value)


def normalize_black_case_no(value):
    raw = str(value or '').strip()
    if not raw:
        return ''

    raw = re.sub(r'\s*/\s*', '/', raw)
    raw = re.sub(r'\s+', ' ', raw)
    match = re.fullmatch('([A-Za-z\u0E01-\u0E2E]{1,8})\\s*([A-Za-z]?\\d{1,8})/(25\\d{2})', raw)
    if not match:
        return None

    return f'{match.group(1)}{match.group(2)}/{match.group(3)}'


def normalize_red_case_no(value):
    return normalize_black_case_no(value)


def parse_float(val, default=0.0):
    try:
        return round(float(str(val).replace(',', '')), 2) if val not in (None, '') else default
    except Exception:
        return default


def parse_required_float(val, field_label):
    if val in (None, ''):
        raise ValueError(f'{field_label}ว่างเปล่า')

    text = str(val).replace(',', '').strip()
    if not text:
        raise ValueError(f'{field_label}ว่างเปล่า')

    try:
        return round(float(text), 2)
    except Exception:
        raise ValueError(f'{field_label}ต้องเป็นตัวเลข')


def parse_required_int(val, field_label):
    if val in (None, ''):
        raise ValueError(f'{field_label}ว่างเปล่า')

    if isinstance(val, int):
        return val
    if isinstance(val, float) and val.is_integer():
        return int(val)

    text = str(val).replace(',', '').strip()
    if not text:
        raise ValueError(f'{field_label}ว่างเปล่า')
    if not re.fullmatch(r'\d+', text):
        raise ValueError(f'{field_label}ต้องเป็นจำนวนเต็ม')

    return int(text)


def parse_required_money(val):
    if val in (None, ''):
        raise ValueError('ว่างเปล่า')

    s = str(val).replace(',', '').strip()

    if not re.fullmatch(r'\d+(\.\d{1,2})?', s):
        raise ValueError('ต้องเป็นตัวเลข และมีทศนิยมได้สูงสุด 2 ตำแหน่ง')

    amount = round(float(s), 2)

    if amount <= 0:
        raise ValueError('ต้องมากกว่า 0')

    return amount

def parse_int(val, default=0):
    try:
        return int(float(str(val).replace(',', ''))) if val not in (None, '') else default
    except Exception:
        return default


def parse_required_positive_int(val):
    if val in (None, ''):
        raise ValueError('ว่างเปล่า')
    if isinstance(val, int):
        value = val
    elif isinstance(val, float) and val.is_integer():
        value = int(val)
    else:
        text = str(val).replace(',', '').strip()
        if not re.fullmatch(r'\d+', text):
            raise ValueError('ต้องเป็นจำนวนเต็มเท่านั้น')
        value = int(text)
    if value <= 0:
        raise ValueError('ต้องมากกว่า 0')
    return value


# ============================================================
# Template Download
# ============================================================

@bp.route('/template/customer', methods=['GET'])
def download_customer_template():
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    from openpyxl.styles import PatternFill, Font, Alignment
    header_fill = PatternFill(start_color='2D3282', end_color='2D3282', fill_type='solid')
    label_fill  = PatternFill(start_color='EEF2FF', end_color='EEF2FF', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    label_font  = Font(bold=True, color='2D3282')

    keys    = [
        'account_no', 'full_name', 'black_case_no', 'filing_date', 'filing_capital',
        'default_date', 'pre_filing_dpd_days', 'filing_note',
    ]
    labels  = [
        'เลขที่บัญชี * (Text 12 หลัก)',
        'ชื่อ-นามสกุล *',
        'คดีหมายเลขดำที่ *',
        'วันที่ยื่นฟ้อง * (YYYY-MM-DD)',
        'ทุนทรัพย์ที่ฟ้อง *',
        'วันที่ผิดนัดชำระก่อนฟ้อง * (YYYY-MM-DD)',
        'DPD ณ วันที่ก่อนส่งฟ้องศาล 1 วัน *',
        'หมายเหตุ / เงื่อนไขพิเศษเพิ่มเติม (ไม่เกิน 100 ตัวอักษร)'
    ]
    example = ['700004761131', 'มานิตย์ บุญรอด', 'ผบE814/2569', '2026-03-12', 250000, '2026-02-10', 30, '']

    ws.append(keys)
    ws.append(labels)
    ws.append(example)

    for row in ws.iter_rows(min_row=1, max_row=5000):
        row[0].number_format = '@'
        row[2].number_format = '@'
        row[3].number_format = '@'
        row[4].number_format = '#,##0.00'
        row[5].number_format = '@'
        row[6].number_format = '0'

    for cell in ws[1]: cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(horizontal='center')
    for cell in ws[2]: cell.fill = label_fill;  cell.font = label_font

    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='Template-Customer_Bulk_Upload.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@bp.route('/template/payment', methods=['GET'])
def download_payment_template():
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    from openpyxl.styles import PatternFill, Font, Alignment
    header_fill = PatternFill(start_color='2D3282', end_color='2D3282', fill_type='solid')
    label_fill  = PatternFill(start_color='EEF2FF', end_color='EEF2FF', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    label_font  = Font(bold=True, color='2D3282')

    keys    = ['account_no', 'payment_date', 'amount']
    labels  = ['เลขที่บัญชี * (Text 12 หลัก)', 'วันที่ชำระเงิน * (YYYY-MM-DD)', 'จำนวนเงิน *']
    example = ['700004761131', '2026-03-12', 30000]

    ws.append(keys)
    ws.append(labels)
    ws.append(example)

    for row in ws.iter_rows(min_row=1, max_row=5000):
        row[0].number_format = '@'
        row[1].number_format = '@'

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for cell in ws[2]:
        cell.fill = label_fill
        cell.font = label_font

    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='Template-Payment_Bulk_Upload.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@bp.route('/template/judgment', methods=['GET'])
def download_judgment_template():
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    from openpyxl.styles import PatternFill, Font, Alignment
    header_fill = PatternFill(start_color='2D3282', end_color='2D3282', fill_type='solid')
    label_fill  = PatternFill(start_color='EEF2FF', end_color='EEF2FF', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    label_font  = Font(bold=True, color='2D3282')

    keys = [
        'account_no', 'judgment_type', 'red_case_no', 'judgment_date', 'total_debt', 'principal_debt',
        'interest_rate', 'court_fee', 'lawyer_fee', 'installment_months', 'first_due_date',
        'step_1_amount', 'step_2_amount', 'step_3_amount', 'step_4_amount', 'default_interest',
        'judgment_note',
    ]
    labels = [
        'เลขที่บัญชี * (Text 12 หลัก)', 'ประเภทคำพิพากษา * (พิพากษาตามยอม/พิพากษาฝ่ายเดียว)',
        'คดีหมายเลขแดงที่ *', 'วันที่พิพากษา * (YYYY-MM-DD)', 'ยอดหนี้รวมตามคำพิพากษา *', 'เงินต้นตามคำพิพากษา *',
        'อัตราดอกเบี้ยต่อปี (%)', 'ค่าธรรมเนียมศาล', 'ค่าทนายความ',
        'จำนวนงวดผ่อนชำระ *', 'วันครบกำหนดงวดแรก * (YYYY-MM-DD)',
        'ค่างวด งวดที่ 1 *', 'ค่างวด งวดที่ 2', 'ค่างวด งวดที่ 3', 'ค่างวด งวดที่ 4',
        'ดอกเบี้ยเมื่อผิดนัด (%)', 'หมายเหตุ / เงื่อนไขพิเศษเพิ่มเติม (ไม่เกิน 100 ตัวอักษร)',
    ]
    example = [
        '700004761131', 'พิพากษาตามยอม', 'ผบ1234/2567', '2026-02-18', 190000, 190000,
        0, 3000, 2000, 24, '2026-03-18',
        8500, 0, 0, 0, 15, '',
    ]

    ws.append(keys)
    ws.append(labels)
    ws.append(example)

    for row in ws.iter_rows(min_row=1, max_row=5000):
        row[0].number_format = '@'
        row[2].number_format = '@'
        row[3].number_format = '@'
        row[10].number_format = '@'

    for cell in ws[1]: cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(horizontal='center')
    for cell in ws[2]: cell.fill = label_fill;  cell.font = label_font

    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='Template-Judgment_Bulk_Upload.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@bp.route('/template/enforcement', methods=['GET'])
def download_enforcement_template():
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    from openpyxl.styles import PatternFill, Font, Alignment
    header_fill = PatternFill(start_color='2D3282', end_color='2D3282', fill_type='solid')
    label_fill  = PatternFill(start_color='EEF2FF', end_color='EEF2FF', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    label_font  = Font(bold=True, color='2D3282')

    keys    = ['account_no', 'enforcement_order_no', 'enforcement_judgment_date', 'enforcement_received_date']
    labels  = [
        'เลขที่บัญชี * (Text 12 หลัก)',
        'เลขหมายบังคับคดี *',
        'วันที่มีคำพิพากษา * (YYYY-MM-DD)',
        'วันที่ได้รับหมาย * (YYYY-MM-DD)'
    ]
    example = ['700004761131', 'EF-2026-00123', '2026-03-18', '2026-03-25']

    ws.append(keys)
    ws.append(labels)
    ws.append(example)

    for row in ws.iter_rows(min_row=1, max_row=5000):
        row[0].number_format = '@'
        row[2].number_format = '@'
        row[3].number_format = '@'

    for cell in ws[1]: cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(horizontal='center')
    for cell in ws[2]: cell.fill = label_fill;  cell.font = label_font

    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='Template-Enforcement_Bulk_Upload.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ============================================================
# Customer Import
# ============================================================

@bp.route('/customer', methods=['POST'])
def import_customers():
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    if 'file' not in request.files:
        return jsonify({'error': 'กรุณาแนบไฟล์ Excel'}), 400

    file = request.files['file']
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': 'รองรับเฉพาะไฟล์ .xlsx'}), 400

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({'error': f'ไม่สามารถอ่านไฟล์ได้: {str(e)}'}), 400

    db = get_db()
    results = []
    success = skip = error = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if all(v is None for v in row):
            continue

        account_no = parse_account_no(row[0])
        if not account_no:
            results.append({'row': row_idx, 'status': 'error', 'message': 'เลขที่บัญชีว่างหรือไม่ถูกต้อง'})
            error += 1
            continue
        if 'e+' in str(account_no).lower() or 'e-' in str(account_no).lower():
            results.append({
                'row': row_idx,
                'account_no': account_no,
                'status': 'error',
                'message': 'เลขที่บัญชีถูก Excel แปลงเป็น Scientific notation กรุณาตั้งค่า column เป็น Text แล้วกรอกเลขบัญชีใหม่'
            })
            error += 1
            continue

        if not account_no.isdigit() or len(account_no) != 12:
            results.append({'row': row_idx, 'account_no': account_no, 'status': 'error',
                            'message': f'เลขที่บัญชีต้องเป็นตัวเลข 12 หลัก'})
            error += 1
            continue

        name = str(row[1]).strip() if row[1] else None
        if not name:
            results.append({'row': row_idx, 'account_no': account_no, 'status': 'error',
                            'message': 'ชื่อ-นามสกุลว่างเปล่า'})
            error += 1
            continue
        if not is_valid_customer_name(name):
            results.append({'row': row_idx, 'account_no': account_no, 'status': 'error',
                            'message': 'ชื่อ-นามสกุล/ชื่อบริษัทใช้ได้เฉพาะตัวอักษร ตัวเลข เว้นวรรค จุด (.) และขีดกลาง (-)'})
            error += 1
            continue

        black_case_no = normalize_black_case_no(row[2] if len(row) > 2 else None)
        if not black_case_no:
            results.append({
                'row': row_idx,
                'account_no': account_no,
                'status': 'error',
                'message': 'คดีหมายเลขดำที่ว่างเปล่าหรือรูปแบบไม่ถูกต้อง เช่น ผบ1234/2567, ผบE814/2569 หรือ พE325/2568'
            })
            error += 1
            continue

        filing_date_raw = row[3] if len(row) > 3 else None
        filing_date = parse_date(filing_date_raw)

        if not filing_date_raw:
            results.append({
                'row': row_idx,
                'account_no': account_no,
                'status': 'error',
                'message': 'วันที่ยื่นฟ้องว่างเปล่า'
            })
            error += 1
            continue

        if not filing_date:
            results.append({
                'row': row_idx,
                'account_no': account_no,
                'status': 'error',
                'message': 'วันที่ยื่นฟ้องต้องเป็นรูปแบบ YYYY-MM-DD เช่น 2026-03-12'
            })
            error += 1
            continue
        
        filing_capital_raw = row[4] if len(row) > 4 else None
        try:
            filing_capital = parse_required_money(filing_capital_raw)
        except ValueError as e:
            results.append({
                'row': row_idx,
                'account_no': account_no,
                'status': 'error',
                'message': f'ทุนทรัพย์ที่ฟ้อง{str(e)}'
            })
            error += 1
            continue

        default_date_raw = row[5] if len(row) > 5 else None
        default_date = parse_date(default_date_raw)
        if not default_date_raw:
            results.append({
                'row': row_idx,
                'account_no': account_no,
                'status': 'error',
                'message': 'วันที่ผิดนัดชำระก่อนฟ้องว่างเปล่า'
            })
            error += 1
            continue
        if not default_date:
            results.append({
                'row': row_idx,
                'account_no': account_no,
                'status': 'error',
                'message': 'วันที่ผิดนัดชำระก่อนฟ้องต้องเป็นรูปแบบ YYYY-MM-DD เช่น 2026-02-10'
            })
            error += 1
            continue

        filing_date_obj = date.fromisoformat(filing_date)
        default_date_obj = date.fromisoformat(default_date)
        today = date.today()
        if filing_date_obj > today:
            results.append({'row': row_idx, 'account_no': account_no, 'status': 'error',
                            'message': 'วันที่ยื่นฟ้องต้องไม่เป็นวันที่ในอนาคต'})
            error += 1
            continue
        if default_date_obj > today:
            results.append({'row': row_idx, 'account_no': account_no, 'status': 'error',
                            'message': 'วันที่ผิดนัดชำระก่อนฟ้องต้องไม่เป็นวันที่ในอนาคต'})
            error += 1
            continue
        if default_date_obj > filing_date_obj:
            results.append({'row': row_idx, 'account_no': account_no, 'status': 'error',
                            'message': 'วันที่ผิดนัดชำระก่อนฟ้องต้องไม่มากกว่าวันที่ยื่นฟ้อง'})
            error += 1
            continue

        try:
            pre_filing_dpd_days = parse_required_positive_int(row[6] if len(row) > 6 else None)
        except ValueError as e:
            results.append({'row': row_idx, 'account_no': account_no, 'status': 'error',
                            'message': f'DPD ณ วันที่ก่อนส่งฟ้องศาล 1 วัน{str(e)}'})
            error += 1
            continue

        filing_note = str(row[7]).strip() if len(row) > 7 and row[7] not in (None, '') else ''
        if len(filing_note) > 100:
            results.append({'row': row_idx, 'account_no': account_no, 'status': 'error',
                            'message': 'หมายเหตุ / เงื่อนไขพิเศษเพิ่มเติมต้องไม่เกิน 100 ตัวอักษร'})
            error += 1
            continue
        
        exists = db.execute('SELECT id, is_deleted FROM customers WHERE account_no = ?', (account_no,)).fetchone()
        if exists:
            msg = 'เลขที่บัญชีนี้เคยถูกลบออกจากระบบแล้ว กรุณากู้คืนข้อมูลก่อน Import' if exists['is_deleted'] == 1 else 'มีข้อมูลในระบบแล้ว'
            results.append({'row': row_idx, 'account_no': account_no, 'status': 'skip', 'message': msg})
            skip += 1
            continue

        try:
            db.execute('''
                INSERT INTO customers (
                    account_no, name, black_case_no, filing_date, filing_capital,
                    default_date, pre_filing_dpd_days, filing_note,
                    case_status, created_by
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'ยื่นฟ้อง', ?)
            ''', (
                account_no, name, black_case_no, filing_date, filing_capital,
                default_date, pre_filing_dpd_days, filing_note or None, user['id']
            ))

            db.execute('''
                INSERT INTO case_status_logs
                (account_no, from_status, to_status, changed_by, note)
                VALUES (?, NULL, 'ยื่นฟ้อง', ?, 'Bulk Import ยื่นฟ้อง')
            ''', (account_no, user['id']))

            results.append({
                'row': row_idx,
                'account_no': account_no,
                'name': name,
                'black_case_no': black_case_no,
                'filing_capital': filing_capital,
                'status': 'success',
                'message': 'นำเข้าสำเร็จ'
            })
            success += 1
        except Exception as e:
            err_str = str(e)
            if 'UNIQUE constraint' in err_str:
                err_str = 'เลขที่บัญชีนี้มีในระบบแล้ว (ซ้ำ)'
            results.append({'row': row_idx, 'account_no': account_no, 'status': 'error', 'message': err_str})
            error += 1

    cur = db.execute(
        """INSERT INTO import_logs (imported_by, import_type, filename, total_rows, success_rows, skip_rows, error_rows, results_json)
           VALUES (?,?,?,?,?,?,?,?)""",
        (user['id'], 'customer_filing', file.filename, success+skip+error, success, skip, error,
         json.dumps(results, ensure_ascii=False))
    )
    db.commit()

    for r in results:
        if r['status'] == 'success':
            refresh_customer_list_cache(r['account_no'], db=db, commit=False)
    db.commit()

    return jsonify({
        'summary': {'success': success, 'skip': skip, 'error': error, 'total': success + skip + error},
        'results': results, 'log_id': cur.lastrowid
    }), 200

@bp.route('/judgment', methods=['POST'])
def import_judgment():
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401
    if user['role'] != 'admin':
        return jsonify({'error': 'เฉพาะ Admin เท่านั้น'}), 403

    if 'file' not in request.files:
        return jsonify({'error': 'กรุณาแนบไฟล์ Excel'}), 400

    file = request.files['file']
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': 'รองรับเฉพาะไฟล์ .xlsx'}), 400

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({'error': f'ไม่สามารถอ่านไฟล์ได้: {str(e)}'}), 400

    db = get_db()
    results = []
    success = skip = error = 0

    VALID_TYPES = ['พิพากษาตามยอม', 'พิพากษาฝ่ายเดียว']

    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if all(v is None for v in row):
            continue

        account_no    = parse_account_no(row[0])
        judgment_type = str(row[1]).strip() if row[1] else None

        if not account_no:
            results.append({'row': row_idx, 'status': 'error', 'message': 'เลขที่บัญชีว่างเปล่า'})
            error += 1
            continue

        if 'e+' in str(account_no).lower() or 'e-' in str(account_no).lower():
            results.append({
                'row': row_idx,
                'account_no': account_no,
                'status': 'error',
                'message': 'เลขที่บัญชีถูก Excel แปลงเป็น Scientific notation กรุณาตั้งค่า column เป็น Text แล้วกรอกเลขบัญชีใหม่'
            })
            error += 1
            continue

        if not account_no.isdigit() or len(account_no) != 12:
            results.append({
                'row': row_idx,
                'account_no': account_no,
                'status': 'error',
                'message': 'เลขที่บัญชีต้องเป็นตัวเลข 12 หลัก'
            })
            error += 1
            continue

        if judgment_type not in VALID_TYPES:
            results.append({'row': row_idx, 'account_no': account_no, 'status': 'error',
                            'message': f'judgment_type ต้องเป็น {" หรือ ".join(VALID_TYPES)}'})
            error += 1
            continue

        cus = db.execute(
            'SELECT * FROM customers WHERE account_no = ? AND is_deleted = 0', (account_no,)
        ).fetchone()
        if not cus:
            results.append({'row': row_idx, 'account_no': account_no, 'status': 'error', 'message': 'ไม่พบในระบบ'})
            error += 1
            continue
        cus = dict(cus)

        if cus['case_status'] != 'ยื่นฟ้อง':
            results.append({'row': row_idx, 'account_no': account_no, 'status': 'skip',
                            'message': f'สถานะปัจจุบันคือ {cus["case_status"]} ไม่ใช่ ยื่นฟ้อง'})
            skip += 1
            continue

        try:
            red_case_no = normalize_red_case_no(row[2] if len(row) > 2 else None)
            if not red_case_no:
                raise ValueError('คดีหมายเลขแดงที่ว่างเปล่าหรือรูปแบบไม่ถูกต้อง เช่น ผบ1234/2567, ผบE814/2569 หรือ พE325/2568')

            judgment_date_raw  = row[3]
            first_due_date_raw = row[10]

            judgment_date  = parse_date(judgment_date_raw)
            first_due_date = parse_date(first_due_date_raw)
            last_due_date  = None

            if not judgment_date:
                raise ValueError('วันที่พิพากษาต้องเป็นรูปแบบ YYYY-MM-DD เช่น 2026-02-18')
            if not first_due_date:
                raise ValueError('วันครบกำหนดงวดแรกต้องเป็นรูปแบบ YYYY-MM-DD เช่น 2026-03-18')
            validate_judgment_timeline(
                cus.get('filing_date'),
                judgment_type,
                judgment_date,
                first_due_date,
            )

            total_debt     = parse_required_float(row[4], 'ยอดหนี้รวม')
            principal      = parse_required_float(row[5], 'เงินต้น')
            interest_rate  = parse_required_float(str(row[6]).replace('%', '').strip() if row[6] not in (None, '') else row[6], 'ดอกเบี้ย')
            court_fee      = parse_float(row[7])
            lawyer_fee     = parse_float(row[8])
            inst_count     = parse_required_int(row[9], 'จำนวนงวด')
            inst1          = parse_required_float(row[11], 'ค่างวด 1')
            inst2          = parse_float(row[12])
            inst3          = parse_float(row[13])
            inst4          = parse_float(row[14])
            default_rate   = parse_required_float(str(row[15]).replace('%', '').strip() if row[15] not in (None, '') else row[15], 'ดอกเบี้ยผิดนัด')
            judgment_note  = str(row[16]).strip() if len(row) > 16 and row[16] not in (None, '') else ''
            if len(judgment_note) > 100:
                raise ValueError('หมายเหตุ / เงื่อนไขพิเศษเพิ่มเติมต้องไม่เกิน 100 ตัวอักษร')

            if not last_due_date and first_due_date and inst_count > 0:
                from datetime import date as _date
                last_due_date = (
                    _date.fromisoformat(first_due_date) + relativedelta(months=inst_count - 1)
                ).isoformat()

            db.execute('''
                UPDATE customers SET
                    case_status           = ?,
                    red_case_no           = ?,
                    judgment_date         = ?,
                    judgment_note         = ?,
                    total_debt            = ?,
                    principal             = ?,
                    interest_rate         = ?,
                    court_fee             = ?,
                    lawyer_fee            = ?,
                    installment_count     = ?,
                    default_interest_rate = ?,
                    first_due_date        = ?,
                    last_due_date         = ?,
                    installment_1         = ?,
                    installment_2         = ?,
                    installment_3         = ?,
                    installment_4         = ?,
                    updated_at            = CURRENT_TIMESTAMP
                WHERE account_no = ? AND is_deleted = 0
            ''', (
                judgment_type, red_case_no, judgment_date, judgment_note or None, total_debt, principal,
                interest_rate, court_fee, lawyer_fee, inst_count, default_rate,
                first_due_date, last_due_date, inst1, inst2, inst3, inst4,
                account_no
            ))

            db.execute('''
                INSERT INTO case_status_logs
                (account_no, from_status, to_status, changed_by, note)
                VALUES (?, ?, ?, ?, 'Bulk Import คำพิพากษา')
            ''', (account_no, cus['case_status'], judgment_type, user['id']))

            results.append({'row': row_idx, 'account_no': account_no,
                            'status': 'success', 'message': f'บันทึกสำเร็จ → {judgment_type}'})
            success += 1

        except Exception as e:
            results.append({'row': row_idx, 'account_no': account_no, 'status': 'error', 'message': str(e)})
            error += 1

    db.commit()

    for r in results:
        if r['status'] == 'success':
            refresh_customer_status(r['account_no'], db)

    cur = db.execute(
        """INSERT INTO import_logs (imported_by, import_type, filename, total_rows, success_rows, skip_rows, error_rows, results_json)
           VALUES (?,?,?,?,?,?,?,?)""",
        (user['id'], 'judgment', file.filename, success+skip+error, success, skip, error,
         json.dumps(results, ensure_ascii=False))
    )
    db.commit()

    return jsonify({
        'summary': {'success': success, 'skip': skip, 'error': error, 'total': success + skip + error},
        'results': results, 'log_id': cur.lastrowid
    }), 200


# ============================================================
# Enforcement Import (admin only)
# ============================================================

@bp.route('/enforcement', methods=['POST'])
def import_enforcement():
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401
    if user['role'] != 'admin':
        return jsonify({'error': 'เฉพาะ Admin เท่านั้น'}), 403

    if 'file' not in request.files:
        return jsonify({'error': 'กรุณาแนบไฟล์ Excel'}), 400

    file = request.files['file']
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': 'รองรับเฉพาะไฟล์ .xlsx'}), 400

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({'error': f'ไม่สามารถอ่านไฟล์ได้: {str(e)}'}), 400

    db = get_db()
    results = []
    success = skip = error = 0

    ALLOWED_STATUS = ['พิพากษาตามยอม', 'พิพากษาฝ่ายเดียว']

    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if all(v is None for v in row):
            continue

        account_no = parse_account_no(row[0])
        if not account_no:
            results.append({'row': row_idx, 'status': 'error', 'message': 'เลขที่บัญชีว่างเปล่า'})
            error += 1
            continue
        if 'e+' in str(account_no).lower() or 'e-' in str(account_no).lower():
            results.append({
                'row': row_idx,
                'account_no': account_no,
                'status': 'error',
                'message': 'เลขที่บัญชีถูก Excel แปลงเป็น Scientific notation กรุณาตั้งค่า column เป็น Text แล้วกรอกเลขบัญชีใหม่'
            })
            error += 1
            continue

        if not account_no.isdigit() or len(account_no) != 12:
            results.append({
                'row': row_idx,
                'account_no': account_no,
                'status': 'error',
                'message': 'เลขที่บัญชีต้องเป็นตัวเลข 12 หลัก'
            })
            error += 1
            continue

        cus = db.execute(
            'SELECT * FROM customers WHERE account_no = ? AND is_deleted = 0', (account_no,)
        ).fetchone()
        if not cus:
            results.append({'row': row_idx, 'account_no': account_no, 'status': 'error', 'message': 'ไม่พบในระบบ'})
            error += 1
            continue
        cus = dict(cus)

        if cus['case_status'] not in ALLOWED_STATUS:
            results.append({'row': row_idx, 'account_no': account_no, 'status': 'skip',
                            'message': f'สถานะปัจจุบันคือ {cus["case_status"]} ไม่สามารถบันทึกหมายได้'})
            skip += 1
            continue

        try:
            enforcement_order_no      = str(row[1]).strip() if row[1] else None
            enforcement_judgment_date = parse_date(row[2])
            enforcement_received_date = parse_date(row[3])

            if not enforcement_order_no:
                raise ValueError('เลขหมายบังคับคดีว่างเปล่า')
            if not enforcement_judgment_date:
                raise ValueError('วันที่มีคำพิพากษาต้องเป็นรูปแบบ YYYY-MM-DD เช่น 2026-03-18')
            if not enforcement_received_date:
                raise ValueError('วันที่ได้รับหมายต้องเป็นรูปแบบ YYYY-MM-DD เช่น 2026-03-25')

            payments = db.execute(
                'SELECT * FROM payments WHERE account_no = ? ORDER BY payment_date ASC', (account_no,)
            ).fetchall()
            payments = [dict(p) for p in payments]

            snap = get_snapshot_at_date(cus, payments, date.today().isoformat())
            if not snap or not (snap['outstanding'] > 0 and snap['ncb_months'] in ['30', '31']):
                results.append({'row': row_idx, 'account_no': account_no, 'status': 'skip',
                                'message': 'ลูกหนี้ยังไม่มีการผิดนัดชำระ ไม่สามารถบันทึกหมายได้'})
                skip += 1
                continue

            db.execute('''
                UPDATE customers SET
                    case_status               = 'บังคับคดี',
                    enforcement_order_no      = ?,
                    enforcement_judgment_date = ?,
                    enforcement_received_date = ?,
                    enforcement_recorded_by   = ?,
                    enforcement_recorded_at   = CURRENT_TIMESTAMP,
                    updated_at                = CURRENT_TIMESTAMP
                WHERE account_no = ? AND is_deleted = 0
            ''', (
                enforcement_order_no, enforcement_judgment_date,
                enforcement_received_date, user['id'], account_no
            ))

            db.execute('''
                INSERT INTO case_status_logs
                (account_no, from_status, to_status, changed_by, note)
                VALUES (?, ?, 'บังคับคดี', ?, 'Bulk Import หมายบังคับคดี')
            ''', (account_no, cus['case_status'], user['id']))

            results.append({'row': row_idx, 'account_no': account_no,
                            'status': 'success', 'message': f'บันทึกหมายบังคับคดีสำเร็จ ({enforcement_order_no})'})
            success += 1

        except Exception as e:
            results.append({'row': row_idx, 'account_no': account_no, 'status': 'error', 'message': str(e)})
            error += 1

    db.commit()

    for r in results:
        if r['status'] == 'success':
            refresh_customer_list_cache(r['account_no'], db=db, commit=False)
    db.commit()

    cur = db.execute(
        """INSERT INTO import_logs (imported_by, import_type, filename, total_rows, success_rows, skip_rows, error_rows, results_json)
           VALUES (?,?,?,?,?,?,?,?)""",
        (user['id'], 'enforcement', file.filename, success+skip+error, success, skip, error,
         json.dumps(results, ensure_ascii=False))
    )
    db.commit()

    return jsonify({
        'summary': {'success': success, 'skip': skip, 'error': error, 'total': success + skip + error},
        'results': results, 'log_id': cur.lastrowid
    }), 200


# ============================================================
# Payment Import
# ============================================================

@bp.route('/payment', methods=['POST'])
def import_payments():
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    if 'file' not in request.files:
        return jsonify({'error': 'กรุณาแนบไฟล์ Excel'}), 400

    file = request.files['file']
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': 'รองรับเฉพาะไฟล์ .xlsx'}), 400

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({'error': f'ไม่สามารถอ่านไฟล์ได้: {str(e)}'}), 400

    db = get_db()
    results = []
    success = skip = error = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if all(v is None for v in row):
            continue

        account_no = parse_account_no(row[0])
        if not account_no:
            results.append({'row': row_idx, 'status': 'error', 'message': 'เลขที่บัญชีว่างเปล่า'})
            error += 1
            continue
        if 'e+' in str(account_no).lower() or 'e-' in str(account_no).lower():
            results.append({
                'row': row_idx,
                'account_no': account_no,
                'status': 'error',
                'message': 'เลขที่บัญชีถูก Excel แปลงเป็น Scientific notation กรุณาตั้งค่า column เป็น Text แล้วกรอกเลขบัญชีใหม่'
            })
            error += 1
            continue
        if not account_no.isdigit():
            results.append({'row': row_idx, 'account_no': account_no, 'status': 'error', 'message': f'เลขที่บัญชีต้องเป็นตัวเลขเท่านั้น: "{account_no}"'})
            error += 1
            continue
        if len(account_no) != 12:
            results.append({'row': row_idx, 'account_no': account_no, 'status': 'error', 'message': f'เลขที่บัญชีต้องมี 12 หลัก (ปัจจุบัน {len(account_no)} หลัก)'})
            error += 1
            continue

        try:
            payment_date = parse_date(row[1])
            amount       = parse_float(row[2])

            if not payment_date:
                raise ValueError('วันที่ชำระต้องเป็นรูปแบบ YYYY-MM-DD เช่น 2026-03-12')
            if amount < 0:
                raise ValueError('จำนวนเงินต้องไม่ติดลบ')

            cus = db.execute('SELECT id FROM customers WHERE account_no = ? AND is_deleted = 0', (account_no,)).fetchone()
            if not cus:
                results.append({'row': row_idx, 'account_no': account_no, 'status': 'error', 'message': 'ไม่พบเลขที่บัญชีในระบบ'})
                error += 1
                continue

            # เช็คซ้ำ (account + payment_date + amount เหมือนกัน)
            dup = db.execute(
                'SELECT id FROM payments WHERE account_no = ? AND payment_date = ? AND amount = ?',
                (account_no, payment_date, amount)
            ).fetchone()
            if dup:
                results.append({'row': row_idx, 'account_no': account_no, 'status': 'skip', 'message': f'มีรายการชำระวันที่ {payment_date} จำนวน {amount} แล้ว'})
                skip += 1
                continue

            db.execute('''
                INSERT INTO payments (customer_id, account_no, payment_date, amount, recorded_by)
                VALUES (?, ?, ?, ?, ?)
            ''', (cus['id'], account_no, payment_date, amount, user['id']))
            results.append({'row': row_idx, 'account_no': account_no, 'status': 'success', 'message': f'นำเข้าสำเร็จ ({payment_date} / {amount:,.2f} บ.)'})
            success += 1
        except Exception as e:
            results.append({'row': row_idx, 'account_no': account_no, 'status': 'error', 'message': str(e)})
            error += 1

    cur = db.execute(
        """INSERT INTO import_logs (imported_by, import_type, filename, total_rows, success_rows, skip_rows, error_rows, results_json)
           VALUES (?,?,?,?,?,?,?,?)""",
        (user['id'], 'payment', file.filename, success+skip+error, success, skip, error, json.dumps(results, ensure_ascii=False))
    )
    db.commit()

    refreshed = set(r['account_no'] for r in results if r['status'] == 'success')
    for account_no in refreshed:
        refresh_customer_status(account_no, db)

    return jsonify({
        'summary': {'success': success, 'skip': skip, 'error': error, 'total': success + skip + error},
        'results': results, 'log_id': cur.lastrowid
    }), 200


# ============================================================
# Import History
# ============================================================

@bp.route('/history', methods=['GET'])
def get_import_history():
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    db   = get_db()
    rows = db.execute(
        """SELECT il.*, u.display_name as imported_by_name
           FROM import_logs il
           LEFT JOIN users u ON u.id = il.imported_by
           ORDER BY il.imported_at DESC
           LIMIT 50"""
    ).fetchall()

    result = []
    for r in rows:
        r = dict(r)
        r.pop('results_json', None)
        result.append(r)

    return jsonify({'history': result}), 200


@bp.route('/history/<int:log_id>/download', methods=['GET'])
def download_import_result(log_id):
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    db  = get_db()
    row = db.execute('SELECT * FROM import_logs WHERE id = ?', (log_id,)).fetchone()
    if not row:
        return jsonify({'error': 'ไม่พบข้อมูล'}), 404

    row = dict(row)
    try:
        results = json.loads(row['results_json'] or '[]')
    except Exception:
        results = []

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Import Result'

    from openpyxl.styles import PatternFill, Font
    header_fill = PatternFill(start_color='2D3282', end_color='2D3282', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    headers = ['แถว', 'เลขที่บัญชี', 'ชื่อ / รายละเอียด', 'สถานะ', 'หมายเหตุ']
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    status_label = {'success': 'สำเร็จ', 'skip': 'ข้ามแล้ว (ซ้ำ)', 'error': 'ผิดพลาด'}
    fill_map = {
        'success': PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid'),
        'skip':    PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid'),
        'error':   PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid'),
    }

    for r in results:
        ws.append([
            r.get('row', ''),
            r.get('account_no', ''),
            r.get('name', '') or r.get('message', ''),
            status_label.get(r.get('status',''), r.get('status','')),
            r.get('message', ''),
        ])
        fill = fill_map.get(r.get('status',''))
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill

    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    import_type = 'Customer' if row['import_type'] == 'customer' else 'Payment'
    filename    = f"ImportResult_{import_type}_{row['imported_at'][:10]}.xlsx"

    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
