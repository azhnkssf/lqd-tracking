import json
import io
import re
import openpyxl
from datetime import date, datetime, timedelta
from openpyxl.styles import PatternFill, Font, Alignment
from flask import Blueprint, current_app, request, jsonify, send_file
from app.database import get_db
from app.services.auth_service import get_user_by_token
from app.services.status_service import refresh_customer_status
from app.services.report_service import (
    RETROACTIVE_ENFORCEMENT_REASON_CODE,
    RETROACTIVE_JUDGMENT_REASON_CODE,
    build_retroactive_alerts,
    build_retroactive_enforcement_alert,
    build_retroactive_judgment_alert,
    get_snapshot_at_date,
)
from app.services.customer_list_cache_service import (
    refresh_all_customer_list_cache,
    refresh_customer_list_cache,
)
from app.services.judgment_service import calculate_judgment_difference, with_judgment_difference
from dateutil.relativedelta import relativedelta

bp = Blueprint('customers', __name__, url_prefix='/api/customers')

CASE_STATUS_TRANSITIONS = {
    'ยื่นฟ้อง'         : ['พิพากษาตามยอม', 'พิพากษาฝ่ายเดียว'],
    'พิพากษาตามยอม'   : ['บังคับคดี'],
    'พิพากษาฝ่ายเดียว': ['บังคับคดี'],
    'บังคับคดี'        : ['ปิดบัญชี'],
}


def get_current_user():
    token = request.cookies.get(current_app.config.get('AUTH_COOKIE_NAME', 'token')) or request.headers.get('Authorization', '').replace('Bearer ', '')
    return get_user_by_token(token)


@bp.before_request
def block_superadmin_from_customer_api():
    user = get_current_user()
    if user and user['role'] == 'superadmin':
        return jsonify({'error': 'Superadmin is limited to user management.'}), 403


def _can_edit_customer_detail(user, case_status):
    if not user or case_status == 'ปิดบัญชี':
        return False
    if case_status == 'ยื่นฟ้อง':
        return user['role'] in ('user', 'admin')
    return user['role'] == 'admin'


def _can_record_enforcement_role(user):
    return bool(user and user['role'] in ('user', 'admin'))


def _log_case_status(db, account_no, from_status, to_status, changed_by, note=''):
    db.execute('''
        INSERT INTO case_status_logs
        (account_no, from_status, to_status, changed_by, note)
        VALUES (?, ?, ?, ?, ?)
    ''', (account_no, from_status, to_status, changed_by, note))


def _attach_case_status_logs(db, cus):
    rows = db.execute(
        'SELECT * FROM case_status_logs WHERE account_no = ? ORDER BY id ASC',
        (cus.get('account_no'),)
    ).fetchall()
    cus['_case_status_logs'] = [dict(r) for r in rows]
    return cus


def _parse_iso_date(value):
    if not value:
        return None
    try:
        return date.fromisoformat(str(value)[:10])
    except Exception:
        return None


def _parse_checker_date_arg(name, default_value):
    value = request.args.get(name, '').strip()
    parsed = _parse_iso_date(value)
    return parsed or default_value


def _checker_user_name(row, prefix=''):
    display_name = row.get(f'{prefix}display_name') if prefix else row.get('display_name')
    username = row.get(f'{prefix}username') if prefix else row.get('username')
    return display_name or username or ''


def _checker_autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 3, 12), 55)


def _checker_style_sheet(ws):
    header_fill = PatternFill(start_color='1E1B4B', end_color='1E1B4B', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    body_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = body_alignment

    ws.freeze_panes = 'A2'
    _checker_autofit(ws)


def _checker_add_sheet(wb, title, headers, rows):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    _checker_style_sheet(ws)
    return ws


def _checker_money(value):
    if value is None or value == '':
        return ''
    try:
        return float(value)
    except (TypeError, ValueError):
        return value


def _checker_json_changes(value):
    try:
        data = json.loads(value or '{}')
    except Exception:
        return []
    if not isinstance(data, dict):
        return []
    rows = []
    for field, change in data.items():
        if isinstance(change, dict):
            rows.append((
                field,
                change.get('label') or field,
                change.get('from'),
                change.get('to'),
            ))
        else:
            rows.append((field, field, '', change))
    return rows


def _validate_judgment_timeline(filing_date, judgment_type, judgment_date, first_due_date):
    if filing_date and judgment_date and judgment_date <= filing_date:
        raise ValueError(
            'วันที่พิพากษาต้องมากกว่าวันที่ยื่นฟ้อง '
            f'(วันที่ยื่นฟ้องในระบบ: {filing_date}, วันที่พิพากษาในไฟล์: {judgment_date})'
        )

    if judgment_date and first_due_date and first_due_date < judgment_date:
        raise ValueError(
            'วันครบกำหนดงวดแรกต้องไม่น้อยกว่าวันที่พิพากษา '
            f'(วันที่พิพากษา: {judgment_date}, วันครบกำหนดงวดแรก: {first_due_date})'
        )


def _normalize_black_case_no(value):
    raw = str(value or '').strip()
    if not raw:
        return ''

    raw = re.sub(r'\s*/\s*', '/', raw)
    raw = re.sub(r'\s+', ' ', raw)
    match = re.fullmatch(r'([A-Za-zก-ฮ]{1,8})(?:\s+([A-Za-z]?\d{1,8})|(\d{1,8}))/(25\d{2})', raw)
    if not match:
        return None

    case_type = match.group(1)
    case_no = match.group(2) or match.group(3)
    year = match.group(4)
    return f'{case_type} {case_no}/{year}'


def _normalize_red_case_no(value):
    return _normalize_black_case_no(value)


def _normalize_black_case_no(value):
    raw = str(value or '').strip()
    if not raw:
        return ''

    raw = re.sub(r'\s*/\s*', '/', raw)
    raw = re.sub(r'\s+', ' ', raw)
    match = re.fullmatch('([A-Za-z\u0E01-\u0E2E]{1,8})\\s*([A-Za-z]?\\d{1,8})/(25\\d{2})', raw)
    if not match:
        return None

    return f'{match.group(1)}{match.group(2)}/{match.group(3)}'


def _normalize_red_case_no(value):
    return _normalize_black_case_no(value)


def _parse_positive_int(value):
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = str(value or '').replace(',', '').strip()
    if not re.fullmatch(r'\d+', text):
        return None
    return int(text)


def _is_valid_customer_name(value):
    text = str(value or '').strip()
    if not text:
        return False
    if re.search(r'\s{2,}', text):
        return False
    return bool(re.fullmatch(r'[A-Za-z0-9ก-ฮะ-์.\-\s]+', text))


def _first_day_next_month(d):
    if not d:
        return None
    if d.month == 12:
        return date(d.year + 1, 1, 1)
    return date(d.year, d.month + 1, 1)


def _is_over_grace_period(cus, today=None):
    """
    Grace period rule:
    - due date อยู่เดือนไหน ให้รอจนหมดเดือนนั้นก่อน
    - วันที่ 1 ของเดือนถัดไปเป็นต้นไป ถึงถือว่าเลย grace period
    """
    today = today or date.today()

    due_date = _parse_iso_date(cus.get('first_due_date'))
    if not due_date:
        return False

    return today >= _first_day_next_month(due_date)


def _payment_status_from_snapshot(cus, snap, payments=None, today=None):
    """
    สถานะสำหรับแสดงผลหน้า UI แบบคำนวณสด

    หมายเหตุ:
    - ไม่ควรเชื่อ payment_status ใน DB ตรง ๆ เพราะสถานะบางอย่างเปลี่ยนตามวันที่
    - เช่น ก่อนถึง first_due_date ต้องเป็น 'ยังไม่ถึงกำหนด'
    """
    today = today or date.today()
    payments = payments or []

    case_status = cus.get('case_status')

    if case_status == 'ปิดบัญชี':
        return 'ชำระครบแล้ว'

    if case_status == 'ยื่นฟ้อง':
        return 'ไม่มีแผนชำระ'

    first_due_date = _parse_iso_date(cus.get('first_due_date'))

    # ถ้ายังไม่ถึงกำหนดงวดแรก และยังไม่มี payment จริง
    # ให้แสดงเป็น "ยังไม่ถึงกำหนด" แทน "จ่ายปกติ"
    if first_due_date and today < first_due_date and len(payments) == 0:
        return 'ยังไม่ถึงกำหนด'

    outstanding = float((snap or {}).get('outstanding') or 0)
    dpd_days = int(float((snap or {}).get('dpd_days') or 0))
    dpd_months = int(float((snap or {}).get('dpd_months') or 0))

    if outstanding > 0 or dpd_days > 0 or dpd_months > 0:
        return 'ค้างชำระ'

    return 'ชำระปกติ'


def _can_record_enforcement(cus, snap, today=None):
    today = today or date.today()

    allowed_status = ['พิพากษาตามยอม', 'พิพากษาฝ่ายเดียว']
    outstanding = float((snap or {}).get('outstanding') or 0)

    has_enforcement_order = bool(
        cus.get('enforcement_order_no') or
        cus.get('enforcement_recorded_at')
    )

    return (
        cus.get('case_status') in allowed_status
        and outstanding > 0
        and not has_enforcement_order
        and _is_over_grace_period(cus, today)
    )


def _decorate_customer_for_ui(cus, payments=None, today=None):
    today = today or date.today()
    payments = payments or []

    try:
        snap = get_snapshot_at_date(cus, payments, today.isoformat())
    except Exception:
        snap = None

    computed_payment_status = _payment_status_from_snapshot(
        cus,
        snap,
        payments=payments,
        today=today
    )

    cus['latest_snapshot'] = snap

    # เก็บค่าเดิมจาก DB ไว้ ไม่ทับ เพื่อให้ตรวจสอบย้อนหลังได้
    cus['db_payment_status'] = cus.get('payment_status')

    # ค่าที่คำนวณสดสำหรับ UI
    cus['computed_payment_status'] = computed_payment_status
    cus['display_payment_status'] = computed_payment_status
    if payments:
        latest_payment = max(payments, key=lambda p: p.get('payment_date') or '')
        cus['last_payment_date'] = latest_payment.get('payment_date')
        cus['last_payment_amount'] = latest_payment.get('amount')
    else:
        cus['last_payment_date'] = None
        cus['last_payment_amount'] = 0

    cus['can_record_enforcement'] = _can_record_enforcement(cus, snap, today)
    cus['has_enforcement_order'] = bool(
        cus.get('enforcement_order_no') or
        cus.get('enforcement_recorded_at')
    )

    return cus


def _csv_values(*names):
    values = []
    for name in names:
        for raw in request.args.getlist(name):
            for item in str(raw or '').split(','):
                item = item.strip()
                if item:
                    values.append(item)
    return values


def _parse_float_arg(name):
    value = request.args.get(name, '').strip()
    if not value:
        return None
    try:
        return float(value.replace(',', ''))
    except (TypeError, ValueError):
        return None


def _fetch_payments_by_account(db, account_nos):
    account_nos = [str(a) for a in account_nos if a]
    if not account_nos:
        return {}

    grouped = {account_no: [] for account_no in account_nos}
    for i in range(0, len(account_nos), 500):
        batch = account_nos[i:i + 500]
        placeholders = ','.join(['?'] * len(batch))
        rows = db.execute(
            f'''
            SELECT *
            FROM payments
            WHERE account_no IN ({placeholders})
            ORDER BY account_no ASC, payment_date ASC
            ''',
            batch
        ).fetchall()

        for row in rows:
            payment = dict(row)
            grouped.setdefault(payment.get('account_no'), []).append(payment)
    return grouped


def _payment_status_filter_values(values):
    aliases = {
        'ชำระปกติ': ['ชำระปกติ'],
        'จ่ายปกติ': ['ชำระปกติ'],
        'ยังไม่เริ่มชำระ': ['ยังไม่ถึงกำหนด'],
        'ยังไม่ถึงกำหนด': ['ยังไม่ถึงกำหนด'],
    }
    normalized = set()
    for value in values:
        for item in aliases.get(value, [value]):
            normalized.add(item)
    return normalized


def _remaining_debt_sql_expr():
    return 'COALESCE(ui_remaining_debt, 0)'


def _decorate_customer_from_list_cache(cus):
    snap = {
        'remaining_debt_raw': cus.get('ui_remaining_debt') or 0,
        'principal_bal': cus.get('ui_principal_bal') or 0,
        'principal_bal_raw': cus.get('ui_principal_bal') or 0,
        'outstanding': cus.get('ui_outstanding') or 0,
        'outstanding_raw': cus.get('ui_outstanding') or 0,
        'dpd_days': cus.get('ui_dpd_days') or 0,
        'dpd_months': cus.get('ui_dpd_months') or 0,
        'oldest_due': cus.get('ui_next_due_date'),
    }
    payment_status = cus.get('ui_payment_status') or cus.get('status') or '-'
    cus['latest_snapshot'] = snap
    cus['db_payment_status'] = cus.get('status')
    cus['computed_payment_status'] = payment_status
    cus['display_payment_status'] = payment_status
    cus['remaining_debt'] = cus.get('ui_remaining_debt') or 0
    cus['last_payment_date'] = cus.get('ui_last_payment_date')
    cus['last_payment_amount'] = cus.get('ui_last_payment_amount') or 0
    cus['can_record_enforcement'] = _can_record_enforcement(cus, snap)
    cus['has_enforcement_order'] = bool(
        cus.get('enforcement_order_no') or
        cus.get('enforcement_recorded_at')
    )
    return cus


@bp.route('', methods=['GET'])
def list_customers():
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    try:
        page = max(1, int(request.args.get('page', 1)))
    except (TypeError, ValueError):
        page = 1
    try:
        per_page = min(100, max(1, int(request.args.get('per_page', 10))))
    except (TypeError, ValueError):
        per_page = 10
    account     = request.args.get('account_no', '').strip()
    name        = request.args.get('name', '').strip()
    case_status = request.args.get('case_status', '').strip()
    case_statuses = _csv_values('case_statuses', 'case_status[]')
    payment_statuses = _payment_status_filter_values(_csv_values('payment_statuses', 'payment_status[]'))
    date_field = request.args.get('date_field', 'due').strip() or 'due'
    date_from = _parse_iso_date(request.args.get('date_from', '').strip())
    date_to = _parse_iso_date(request.args.get('date_to', '').strip())
    outstanding_min = _parse_float_arg('outstanding_min')
    outstanding_max = _parse_float_arg('outstanding_max')
    sort_by     = request.args.get('sort_by', '').strip()
    sort_dir    = request.args.get('sort_dir', 'desc').strip().lower()
    offset      = (page - 1) * per_page
    allowed_date_fields = {'due', 'nextDue', 'filingDate', 'judgmentDate', 'enforcementJudgmentDate', 'lastPaymentDate'}
    if date_field not in allowed_date_fields:
        date_field = 'due'

    db = get_db()

    where  = ['is_deleted = 0']
    params = []

    if account:
        where.append('account_no LIKE ?')
        params.append(f'%{account}%')
    if name:
        where.append('name LIKE ?')
        params.append(f'%{name}%')
    if case_status:
        where.append('case_status = ?')
        params.append(case_status)
    if case_statuses:
        placeholders = ','.join(['?'] * len(case_statuses))
        where.append(f'case_status IN ({placeholders})')
        params.extend(case_statuses)
    if payment_statuses:
        placeholders = ','.join(['?'] * len(payment_statuses))
        where.append(f'ui_payment_status IN ({placeholders})')
        params.extend(sorted(payment_statuses))
    if outstanding_min is not None or outstanding_max is not None:
        remaining_debt_expr = _remaining_debt_sql_expr()
        if outstanding_min is not None:
            where.append(f'({remaining_debt_expr}) >= ?')
            params.append(outstanding_min)
        if outstanding_max is not None:
            where.append(f'({remaining_debt_expr}) <= ?')
            params.append(outstanding_max)

    sql_date_columns = {
        'due': 'first_due_date',
        'nextDue': 'ui_next_due_date',
        'filingDate': 'filing_date',
        'judgmentDate': 'judgment_date',
        'enforcementJudgmentDate': 'enforcement_judgment_date',
    }
    if (date_from or date_to) and date_field in sql_date_columns:
        date_column = sql_date_columns[date_field]
        if date_from:
            where.append(f'{date_column} >= ?')
            params.append(date_from.isoformat())
        if date_to:
            where.append(f'{date_column} <= ?')
            params.append(date_to.isoformat())
    elif (date_from or date_to) and date_field == 'lastPaymentDate':
        if date_from:
            where.append('ui_last_payment_date >= ?')
            params.append(date_from.isoformat())
        if date_to:
            where.append('ui_last_payment_date <= ?')
            params.append(date_to.isoformat())

    where_clause = 'WHERE ' + ' AND '.join(where)
    sort_columns = {
        'account_no': 'account_no',
        'filing_date': 'filing_date',
        'judgment_date': 'judgment_date',
    }
    if sort_by in sort_columns:
        direction = 'ASC' if sort_dir == 'asc' else 'DESC'
        order_clause = f'ORDER BY {sort_columns[sort_by]} {direction}, id DESC'
    else:
        order_clause = 'ORDER BY id DESC'

    summary = db.execute('''
        SELECT
            COALESCE(SUM(
                CASE
                    WHEN case_status = 'ยื่นฟ้อง' THEN COALESCE(filing_capital, 0)
                    ELSE COALESCE(total_debt, 0)
                END
            ), 0) as total_value,
            COUNT(*) as active_count
        FROM customers
        WHERE COALESCE(case_status, '') != 'ปิดบัญชี'
          AND is_deleted = 0
    ''').fetchone()

    status_rows = db.execute(
        'SELECT case_status, COUNT(*) as count FROM customers WHERE is_deleted = 0 GROUP BY case_status'
    ).fetchall()
    case_counts = {r['case_status'] or '': r['count'] for r in status_rows}
    case_counts['ทั้งหมด'] = sum(case_counts.values())

    decorated_rows = []

    total = db.execute(
        f'SELECT COUNT(*) FROM customers {where_clause}', params
    ).fetchone()[0]

    rows = db.execute(
        f'SELECT * FROM customers {where_clause} {order_clause} LIMIT ? OFFSET ?',
        params + [per_page, offset]
    ).fetchall()
    page_rows = [dict(r) for r in rows]

    for cus in page_rows:
        decorated_rows.append(_decorate_customer_from_list_cache(cus))

    return jsonify({
        'data':     decorated_rows,
        'total':    total,
        'page':     page,
        'per_page': per_page,
        'summary': {
            'total_value':  summary['total_value'],
            'active_count': summary['active_count'],
            'case_counts':  case_counts
        }
    }), 200


@bp.route('/checker-export', methods=['GET'])
def export_checker_raw():
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401
    if user['role'] != 'admin':
        return jsonify({'error': 'Forbidden'}), 403

    today = date.today()
    default_from = today.replace(day=1)
    date_from = _parse_checker_date_arg('date_from', default_from)
    date_to = _parse_checker_date_arg('date_to', today)
    if date_to < date_from:
        return jsonify({'error': 'date_to must be after date_from'}), 400
    if date_from > today or date_to > today:
        return jsonify({'error': 'future dates are not allowed'}), 400

    start_at = datetime.combine(date_from, datetime.min.time()).isoformat(sep=' ')
    end_at = datetime.combine(date_to + timedelta(days=1), datetime.min.time()).isoformat(sep=' ')

    db = get_db()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    customer_cols = [
        'account_no', 'name', 'case_status', 'status',
        'black_case_no', 'red_case_no',
        'filing_date', 'filing_capital', 'default_date', 'pre_filing_dpd_days',
        'judgment_date', 'total_debt', 'principal', 'judgment_difference',
        'interest_rate', 'court_fee', 'lawyer_fee',
        'installment_count', 'first_due_date', 'last_due_date',
        'installment_1', 'installment_2', 'installment_3', 'installment_4',
        'enforcement_order_no', 'enforcement_judgment_date',
        'enforcement_received_date', 'enforcement_recorded_at',
        'created_at', 'updated_at',
    ]
    money_cols = {
        'filing_capital', 'total_debt', 'principal', 'judgment_difference',
        'interest_rate', 'court_fee', 'lawyer_fee', 'installment_1',
        'installment_2', 'installment_3', 'installment_4',
    }
    customer_value = lambda row, col: _checker_money(row.get(col)) if col in money_cols else row.get(col)

    new_customers = [dict(r) for r in db.execute(f'''
        SELECT c.*, u.display_name, u.username
        FROM customers c
        LEFT JOIN users u ON u.id = c.created_by
        WHERE c.created_at >= ? AND c.created_at < ?
        ORDER BY c.created_at ASC, c.id ASC
    ''', (start_at, end_at)).fetchall()]

    status_logs = [dict(r) for r in db.execute('''
        SELECT csl.*, c.name, c.black_case_no, c.red_case_no,
               u.display_name, u.username
        FROM case_status_logs csl
        LEFT JOIN customers c ON c.account_no = csl.account_no
        LEFT JOIN users u ON u.id = csl.changed_by
        WHERE csl.changed_at >= ? AND csl.changed_at < ?
        ORDER BY csl.changed_at ASC, csl.id ASC
    ''', (start_at, end_at)).fetchall()]

    edits = [dict(r) for r in db.execute('''
        SELECT ce.*, c.name, c.case_status, u.display_name, u.username
        FROM customer_edits ce
        LEFT JOIN customers c ON c.account_no = ce.account_no
        LEFT JOIN users u ON u.id = ce.edited_by
        WHERE ce.edited_at >= ? AND ce.edited_at < ?
        ORDER BY ce.edited_at ASC, ce.id ASC
    ''', (start_at, end_at)).fetchall()]

    deleted_customers = [dict(r) for r in db.execute(f'''
        SELECT c.*, u.display_name, u.username
        FROM customers c
        LEFT JOIN users u ON u.id = c.created_by
        WHERE c.deleted_at >= ? AND c.deleted_at < ?
        ORDER BY c.deleted_at ASC, c.id ASC
    ''', (start_at, end_at)).fetchall()]

    touched_accounts = sorted({
        *(r.get('account_no') for r in new_customers),
        *(r.get('account_no') for r in status_logs),
        *(r.get('account_no') for r in edits),
        *(r.get('account_no') for r in deleted_customers),
    } - {None, ''})

    current_rows = []
    if touched_accounts:
        placeholders = ','.join(['?'] * len(touched_accounts))
        current_rows = [dict(r) for r in db.execute(f'''
            SELECT c.*, creator.display_name AS creator_display_name, creator.username AS creator_username,
                   enf.display_name AS enforcement_display_name, enf.username AS enforcement_username
            FROM customers c
            LEFT JOIN users creator ON creator.id = c.created_by
            LEFT JOIN users enf ON enf.id = c.enforcement_recorded_by
            WHERE c.account_no IN ({placeholders})
            ORDER BY c.account_no ASC
        ''', touched_accounts).fetchall()]

    judgment_statuses = {'พิพากษาตามยอม', 'พิพากษาฝ่ายเดียว'}
    judgment_logs = [r for r in status_logs if r.get('to_status') in judgment_statuses]
    enforcement_logs = [r for r in status_logs if r.get('to_status') == 'บังคับคดี']
    judgment_rows = [dict(r) for r in db.execute('''
        SELECT csl.changed_at, csl.to_status, csl.note, csl.account_no,
               c.name, c.judgment_date, c.red_case_no, c.total_debt, c.principal,
               c.first_due_date, c.installment_count,
               u.display_name, u.username
        FROM case_status_logs csl
        LEFT JOIN customers c ON c.account_no = csl.account_no
        LEFT JOIN users u ON u.id = csl.changed_by
        WHERE csl.changed_at >= ? AND csl.changed_at < ?
          AND csl.to_status IN (?, ?)
        ORDER BY csl.changed_at ASC, csl.id ASC
    ''', (start_at, end_at, 'พิพากษาตามยอม', 'พิพากษาฝ่ายเดียว')).fetchall()]
    enforcement_rows = [dict(r) for r in db.execute('''
        SELECT csl.changed_at, csl.from_status, csl.to_status, csl.note, csl.account_no,
               c.name, c.enforcement_order_no, c.enforcement_judgment_date,
               c.enforcement_received_date, u.display_name, u.username
        FROM case_status_logs csl
        LEFT JOIN customers c ON c.account_no = csl.account_no
        LEFT JOIN users u ON u.id = csl.changed_by
        WHERE csl.changed_at >= ? AND csl.changed_at < ?
          AND csl.to_status = ?
        ORDER BY csl.changed_at ASC, csl.id ASC
    ''', (start_at, end_at, 'บังคับคดี')).fetchall()]

    summary_rows = [
        ['Date From', date_from.isoformat()],
        ['Date To', date_to.isoformat()],
        ['Exported At', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ['Exported By', _checker_user_name(dict(user))],
        ['New Customers', len(new_customers)],
        ['Status Changes', len(status_logs)],
        ['Judgment Recorded', len(judgment_logs)],
        ['Enforcement Recorded', len(enforcement_logs)],
        ['Field Changes', sum(len(_checker_json_changes(r.get('changes'))) for r in edits)],
        ['Deleted Customers', len(deleted_customers)],
        ['Touched Accounts', len(touched_accounts)],
    ]
    _checker_add_sheet(wb, 'Summary', ['Metric', 'Value'], summary_rows)

    _checker_add_sheet(
        wb,
        'New Customers',
        ['created_at', 'maker', *customer_cols],
        [[r.get('created_at'), _checker_user_name(r), *[customer_value(r, col) for col in customer_cols]] for r in new_customers]
    )

    _checker_add_sheet(
        wb,
        'Status Changes',
        ['changed_at', 'maker', 'account_no', 'name', 'from_status', 'to_status', 'note', 'black_case_no', 'red_case_no'],
        [[
            r.get('changed_at'), _checker_user_name(r), r.get('account_no'), r.get('name'),
            r.get('from_status'), r.get('to_status'), r.get('note'),
            r.get('black_case_no'), r.get('red_case_no'),
        ] for r in status_logs]
    )

    _checker_add_sheet(
        wb,
        'Judgment Recorded',
        ['recorded_at', 'maker', 'account_no', 'name', 'judgment_type', 'judgment_date', 'red_case_no', 'total_debt', 'principal', 'first_due_date', 'installment_count', 'note'],
        [[
            r.get('changed_at'), _checker_user_name(r), r.get('account_no'), r.get('name'),
            r.get('to_status'), r.get('judgment_date'), r.get('red_case_no'),
            _checker_money(r.get('total_debt')), _checker_money(r.get('principal')),
            r.get('first_due_date'), r.get('installment_count'), r.get('note'),
        ] for r in judgment_rows]
    )

    _checker_add_sheet(
        wb,
        'Enforcement Recorded',
        ['recorded_at', 'maker', 'account_no', 'name', 'from_status', 'to_status', 'enforcement_order_no', 'enforcement_judgment_date', 'enforcement_received_date', 'note'],
        [[
            r.get('changed_at'), _checker_user_name(r), r.get('account_no'), r.get('name'),
            r.get('from_status'), r.get('to_status'), r.get('enforcement_order_no'),
            r.get('enforcement_judgment_date'), r.get('enforcement_received_date'), r.get('note'),
        ] for r in enforcement_rows]
    )

    field_change_rows = []
    for edit in edits:
        for field, label, old_value, new_value in _checker_json_changes(edit.get('changes')):
            field_change_rows.append([
                edit.get('edited_at'), _checker_user_name(edit), edit.get('account_no'),
                edit.get('name'), edit.get('case_status'), field, label, old_value, new_value,
            ])
    _checker_add_sheet(
        wb,
        'Field Changes',
        ['edited_at', 'maker', 'account_no', 'name', 'current_case_status', 'field', 'label', 'old_value', 'new_value'],
        field_change_rows
    )

    _checker_add_sheet(
        wb,
        'Deleted Customers',
        ['deleted_at', 'created_by', *customer_cols],
        [[r.get('deleted_at'), _checker_user_name(r), *[customer_value(r, col) for col in customer_cols]] for r in deleted_customers]
    )

    _checker_add_sheet(
        wb,
        'Current Snapshot',
        ['created_by', 'enforcement_recorded_by', *customer_cols],
        [[
            _checker_user_name(r, 'creator_'),
            _checker_user_name(r, 'enforcement_'),
            *[customer_value(r, col) for col in customer_cols],
        ] for r in current_rows]
    )

    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = '#,##0.00'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'Checker_Raw_{date_from.isoformat()}_to_{date_to.isoformat()}.xlsx'
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@bp.route('/cache/refresh-all', methods=['POST'])
def refresh_customer_list_cache_all():
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401
    if user['role'] not in ('user', 'admin'):
        return jsonify({'error': 'Forbidden'}), 403

    db = get_db()
    result = refresh_all_customer_list_cache(db=db)
    return jsonify({'message': 'refresh customer list cache สำเร็จ', **result}), 200


@bp.route('/<account_no>/cache/refresh', methods=['POST'])
def refresh_customer_list_cache_one(account_no):
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401
    if user['role'] != 'admin':
        return jsonify({'error': 'เฉพาะ Admin เท่านั้น'}), 403

    db = get_db()
    ok = refresh_customer_list_cache(account_no, db=db)
    if not ok:
        return jsonify({'error': 'ไม่พบข้อมูล'}), 404
    return jsonify({'message': 'refresh customer list cache สำเร็จ', 'account_no': account_no}), 200


@bp.route('', methods=['POST'])
def create_customer():
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    data = request.get_json() or {}

    required = [
        'account_no', 'name', 'black_case_no', 'filing_date', 'filing_capital',
        'default_date', 'pre_filing_dpd_days',
    ]
    missing = [f for f in required if data.get(f) is None or data.get(f) == '']
    if missing:
        field_labels = {
            'account_no': 'หมายเลขบัญชี',
            'name': 'ชื่อ-นามสกุล',
            'black_case_no': 'คดีหมายเลขดำที่',
            'filing_date': 'วันที่ยื่นฟ้อง',
            'filing_capital': 'ทุนทรัพย์ที่ฟ้อง',
            'default_date': 'วันที่ผิดนัดชำระก่อนฟ้อง',
            'pre_filing_dpd_days': 'DPD ณ วันที่ก่อนส่งฟ้องศาล 1 วัน',
        }
        missing_labels = [field_labels.get(f, f) for f in missing]
        return jsonify({'error': f'กรุณากรอกข้อมูลให้ครบ: {", ".join(missing_labels)}'}), 400

    account_no = str(data.get('account_no') or '').strip()
    black_case_no = _normalize_black_case_no(data.get('black_case_no'))
    name = str(data.get('name') or '').strip()
    filing_date = str(data.get('filing_date') or '').strip()
    filing_capital_text = str(data.get('filing_capital') or '').replace(',', '').strip()
    default_date = str(data.get('default_date') or '').strip()
    filing_note = str(data.get('filing_note') or '').strip()

    if not account_no.isdigit() or len(account_no) != 12:
        return jsonify({'error': 'เลขที่บัญชีต้องเป็นตัวเลข 12 หลัก'}), 400

    if not _is_valid_customer_name(name):
        return jsonify({'error': 'ชื่อ-นามสกุล/ชื่อบริษัทใช้ได้เฉพาะตัวอักษร ตัวเลข เว้นวรรค จุด (.) และขีดกลาง (-)'}), 400

    if not re.fullmatch(r'\d{4}-\d{2}-\d{2}', filing_date):
        return jsonify({'error': 'วันที่ยื่นฟ้องต้องเป็นรูปแบบ YYYY-MM-DD'}), 400

    if not black_case_no:
        return jsonify({'error': 'คดีหมายเลขดำที่ต้องอยู่ในรูปแบบ: ตัวย่อประเภทคดีติดเลขที่ฟ้อง/ปี พ.ศ. เช่น ผบ1234/2567, ผบE814/2569 หรือ พE325/2568'}), 400

    if not re.fullmatch(r'\d{4}-\d{2}-\d{2}', default_date):
        return jsonify({'error': 'วันที่ผิดนัดชำระก่อนฟ้องต้องเป็นรูปแบบ YYYY-MM-DD'}), 400

    filing_date_obj = _parse_iso_date(filing_date)
    default_date_obj = _parse_iso_date(default_date)
    today = date.today()
    if not filing_date_obj or filing_date_obj > today:
        return jsonify({'error': 'วันที่ยื่นฟ้องต้องไม่เป็นวันที่ในอนาคต'}), 400
    if not default_date_obj or default_date_obj > today:
        return jsonify({'error': 'วันที่ผิดนัดชำระก่อนฟ้องต้องไม่เป็นวันที่ในอนาคต'}), 400
    if default_date_obj > filing_date_obj:
        return jsonify({'error': 'วันที่ผิดนัดชำระก่อนฟ้องต้องไม่มากกว่าวันที่ยื่นฟ้อง'}), 400

    pre_filing_dpd_days = _parse_positive_int(data.get('pre_filing_dpd_days'))
    if pre_filing_dpd_days is None:
        return jsonify({'error': 'DPD ณ วันที่ก่อนส่งฟ้องศาล 1 วันต้องเป็นจำนวนเต็มเท่านั้น'}), 400
    if pre_filing_dpd_days < 0:
        return jsonify({'error': 'DPD ณ วันที่ก่อนส่งฟ้องศาล 1 วันต้องไม่น้อยกว่า 0'}), 400

    if len(filing_note) > 100:
        return jsonify({'error': 'หมายเหตุ / เงื่อนไขพิเศษเพิ่มเติมต้องไม่เกิน 100 ตัวอักษร'}), 400

    if not re.fullmatch(r'\d+(\.\d{1,2})?', filing_capital_text):
        return jsonify({'error': 'ทุนทรัพย์ที่ฟ้องต้องเป็นตัวเลข และมีทศนิยมได้สูงสุด 2 ตำแหน่ง'}), 400

    filing_capital = round(float(filing_capital_text), 2)
    if filing_capital <= 0:
        return jsonify({'error': 'ทุนทรัพย์ที่ฟ้องต้องมากกว่า 0'}), 400

    db = get_db()

    existing = db.execute(
        'SELECT id FROM customers WHERE account_no = ?', (account_no,)
    ).fetchone()
    if existing:
        return jsonify({'error': 'เลขที่บัญชีนี้มีในระบบแล้ว'}), 409

    db.execute('''
        INSERT INTO customers (
            account_no, name, filing_date, black_case_no, filing_capital,
            default_date, pre_filing_dpd_days, filing_note,
            case_status, created_by
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'ยื่นฟ้อง', ?)
    ''', (
        account_no,
        name,
        filing_date,
        black_case_no,
        filing_capital,
        default_date,
        pre_filing_dpd_days,
        filing_note or None,
        user['id']
    ))

    _log_case_status(db, account_no, None, 'ยื่นฟ้อง', user['id'], 'สร้างข้อมูลใหม่')
    refresh_customer_list_cache(account_no, db=db, commit=False)
    db.commit()

    return jsonify({
        'message': 'บันทึกข้อมูลสำเร็จ',
        'account_no': account_no,
        'black_case_no': black_case_no,
        'filing_capital': filing_capital,
        'default_date': default_date,
        'pre_filing_dpd_days': pre_filing_dpd_days,
        'filing_note': filing_note,
    }), 201


@bp.route('/<account_no>', methods=['GET'])
def get_customer(account_no):
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    db  = get_db()
    row = db.execute(
        'SELECT * FROM customers WHERE account_no = ? AND is_deleted = 0', (account_no,)
    ).fetchone()

    if not row:
        return jsonify({'error': 'ไม่พบข้อมูล'}), 404

    cus = _attach_case_status_logs(db, with_judgment_difference(dict(row)))
    cus['retroactive_enforcement_alert'] = build_retroactive_enforcement_alert(
        cus,
        db=db,
        include_marked=True,
    )
    cus['retroactive_judgment_alert'] = build_retroactive_judgment_alert(
        cus,
        db=db,
        include_marked=True,
    )
    cus['retroactive_alerts'] = build_retroactive_alerts(
        cus,
        db=db,
        include_marked=True,
    )

    payments = db.execute(
        'SELECT * FROM payments WHERE account_no = ? ORDER BY payment_date ASC',
        (account_no,)
    ).fetchall()
    payments = [dict(p) for p in payments]

    cus = _decorate_customer_for_ui(cus, payments)
    cus.pop('_case_status_logs', None)

    return jsonify(cus), 200


@bp.route('/<account_no>/retroactive-enforcement-fix', methods=['POST'])
def mark_retroactive_enforcement_fix(account_no):
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401
    if user['role'] != 'admin':
        return jsonify({'error': 'เฉพาะ Admin เท่านั้นที่ยืนยันการแก้รายงานย้อนหลังได้'}), 403

    db = get_db()
    row = db.execute(
        'SELECT * FROM customers WHERE account_no = ? AND is_deleted = 0',
        (account_no,)
    ).fetchone()
    if not row:
        return jsonify({'error': 'ไม่พบข้อมูล'}), 404

    cus = _attach_case_status_logs(db, dict(row))
    alert = build_retroactive_enforcement_alert(cus, db=db, include_marked=True)
    if not alert:
        return jsonify({'error': 'บัญชีนี้ไม่มีรายการรายงานย้อนหลังที่ต้องยืนยัน'}), 400

    if alert.get('marked'):
        return jsonify({
            'message': 'รายการนี้ถูกยืนยันแล้ว',
            'retroactive_enforcement_alert': alert,
        }), 200

    data = request.get_json(silent=True) or {}
    note = str(data.get('note') or '').strip() or 'แก้รายงานย้อนหลังแล้ว'

    try:
        db.execute('''
            INSERT INTO report_retroactive_fix_marks
            (account_no, affected_report_month, effective_date, reason_code,
             source_report_month, marked_by, note)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (
            account_no,
            alert['affected_report_month'],
            alert['effective_date'],
            RETROACTIVE_ENFORCEMENT_REASON_CODE,
            alert.get('source_report_month'),
            user['id'],
            note,
        ))
        db.execute(
            'INSERT INTO customer_edits (customer_id, account_no, edited_by, changes) VALUES (?, ?, ?, ?)',
            (
                row['id'],
                account_no,
                user['id'],
                json.dumps({
                    'retroactive_enforcement_fix': {
                        'label': 'การแก้รายงานย้อนหลัง',
                        'from': 'รอยืนยัน',
                        'to': f'ยืนยันแล้ว ({alert["affected_month_label"]})',
                    },
                    'retroactive_enforcement_effective_date': {
                        'label': 'วันที่ของหมายบังคับคดี',
                        'from': None,
                        'to': alert['effective_date'],
                    },
                }, ensure_ascii=False),
            )
        )
        db.commit()
    except Exception:
        db.rollback()
        return jsonify({'error': 'ไม่สามารถบันทึกการยืนยันได้ หรือรายการนี้ถูกยืนยันแล้ว'}), 409

    cus = _attach_case_status_logs(db, dict(row))
    updated_alert = build_retroactive_enforcement_alert(cus, db=db, include_marked=True)

    return jsonify({
        'message': f'ยืนยันว่าแก้รายงานเดือน {alert["affected_month_label"]} แล้ว',
        'retroactive_enforcement_alert': updated_alert,
    }), 200


@bp.route('/<account_no>/retroactive-report-fix', methods=['POST'])
def mark_retroactive_report_fix(account_no):
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401
    if user['role'] != 'admin':
        return jsonify({'error': 'เฉพาะ Admin เท่านั้นที่ยืนยันการแก้รายงานย้อนหลังได้'}), 403

    data = request.get_json(silent=True) or {}
    reason_code = str(data.get('reason_code') or '').strip()
    affected_report_month = str(data.get('affected_report_month') or '').strip()
    note = str(data.get('note') or '').strip() or 'แก้รายงานย้อนหลังแล้ว'
    allowed_reason_codes = {
        RETROACTIVE_ENFORCEMENT_REASON_CODE,
        RETROACTIVE_JUDGMENT_REASON_CODE,
    }
    if reason_code not in allowed_reason_codes:
        return jsonify({'error': 'reason_code ไม่ถูกต้อง'}), 400

    db = get_db()
    row = db.execute(
        'SELECT * FROM customers WHERE account_no = ? AND is_deleted = 0',
        (account_no,)
    ).fetchone()
    if not row:
        return jsonify({'error': 'ไม่พบข้อมูล'}), 404

    cus = _attach_case_status_logs(db, dict(row))
    alerts = build_retroactive_alerts(cus, db=db, include_marked=True)
    alert = next((
        item for item in alerts
        if item.get('reason_code') == reason_code
        and (not affected_report_month or item.get('affected_report_month') == affected_report_month)
    ), None)
    if not alert:
        return jsonify({'error': 'บัญชีนี้ไม่มีรายการรายงานย้อนหลังที่ต้องยืนยัน'}), 400

    try:
        existing_mark = db.execute('''
            SELECT id
            FROM report_retroactive_fix_marks
            WHERE account_no = ?
              AND affected_report_month = ?
              AND reason_code = ?
            LIMIT 1
        ''', (
            account_no,
            alert['affected_report_month'],
            reason_code,
        )).fetchone()

        if existing_mark:
            db.execute('''
                UPDATE report_retroactive_fix_marks
                SET effective_date = ?,
                    source_report_month = ?,
                    marked_by = ?,
                    marked_at = CURRENT_TIMESTAMP,
                    note = ?
                WHERE id = ?
            ''', (
                alert['effective_date'],
                alert.get('source_report_month'),
                user['id'],
                note,
                existing_mark['id'],
            ))
        else:
            db.execute('''
                INSERT INTO report_retroactive_fix_marks
                (account_no, affected_report_month, effective_date, reason_code,
                 source_report_month, marked_by, note)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (
                account_no,
                alert['affected_report_month'],
                alert['effective_date'],
                reason_code,
                alert.get('source_report_month'),
                user['id'],
                note,
            ))

        try:
            db.execute(
                'INSERT INTO customer_edits (customer_id, account_no, edited_by, changes) VALUES (?, ?, ?, ?)',
                (
                    row['id'],
                    account_no,
                    user['id'],
                    json.dumps({
                        'retroactive_report_fix': {
                            'label': 'การแก้รายงานย้อนหลัง',
                            'from': 'รอยืนยัน',
                            'to': f'ยืนยันแล้ว ({alert["affected_month_label"]})',
                        },
                        'retroactive_report_fix_reason': {
                            'label': 'ประเภทการแก้ย้อนหลัง',
                            'from': None,
                            'to': reason_code,
                        },
                        'retroactive_report_effective_date': {
                            'label': 'วันที่มีผล',
                            'from': None,
                            'to': alert['effective_date'],
                        },
                    }, ensure_ascii=False),
                )
            )
        except Exception:
            pass
        db.commit()
    except Exception:
        db.rollback()
        return jsonify({'error': 'ไม่สามารถบันทึกการยืนยันได้'}), 409

    cus = _attach_case_status_logs(db, dict(row))
    updated_alerts = build_retroactive_alerts(cus, db=db, include_marked=True)
    updated_alert = next((
        item for item in updated_alerts
        if item.get('reason_code') == reason_code
        and item.get('affected_report_month') == alert['affected_report_month']
    ), None)

    return jsonify({
        'message': f'ยืนยันว่าแก้รายงานเดือน {alert["affected_month_label"]} แล้ว',
        'retroactive_alert': updated_alert,
        'retroactive_alerts': updated_alerts,
        'retroactive_enforcement_alert': next((item for item in updated_alerts if item.get('type') == 'enforcement'), None),
        'retroactive_judgment_alert': next((item for item in updated_alerts if item.get('type') == 'judgment'), None),
    }), 200


@bp.route('/<account_no>/judgment', methods=['PATCH'])
def update_judgment(account_no):
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    db  = get_db()
    cus = db.execute(
        'SELECT * FROM customers WHERE account_no = ? AND is_deleted = 0', (account_no,)
    ).fetchone()
    if not cus:
        return jsonify({'error': 'ไม่พบข้อมูล'}), 404
    cus = dict(cus)

    if not _can_edit_customer_detail(user, cus.get('case_status')):
        return jsonify({'error': 'ไม่มีสิทธิ์แก้ไขข้อมูลในสถานะนี้'}), 403

    data = request.get_json()

    required = [
        'judgment_type', 'red_case_no', 'judgment_date', 'total_debt', 'principal',
        'interest_rate', 'installment_count', 'first_due_date', 'installment_1',
        'default_interest_rate',
    ]
    missing = [f for f in required if data.get(f) is None or data.get(f) == '']
    if missing:
        return jsonify({'error': f'กรุณากรอกข้อมูลให้ครบ: {", ".join(missing)}'}), 400

    allowed = CASE_STATUS_TRANSITIONS.get('ยื่นฟ้อง', [])
    if cus['case_status'] != 'ยื่นฟ้อง':
        return jsonify({'error': f'ไม่สามารถบันทึกคำพิพากษาได้ สถานะปัจจุบันคือ {cus["case_status"]}'}), 400
    if data['judgment_type'] not in allowed:
        return jsonify({'error': f'judgment_type ต้องเป็น {" หรือ ".join(allowed)}'}), 400

    filing_date    = cus.get('filing_date', '')
    red_case_no    = _normalize_red_case_no(data.get('red_case_no'))
    judgment_note  = str(data.get('judgment_note') or '').strip()
    judgment_date  = data.get('judgment_date', '')
    first_due_date = data.get('first_due_date', '')

    if not red_case_no:
        return jsonify({'error': 'คดีหมายเลขแดงที่ต้องอยู่ในรูปแบบ: ตัวย่อประเภทคดีติดเลขที่/ปี พ.ศ. เช่น ผบ1234/2567, ผบE814/2569 หรือ พE325/2568'}), 400
    if len(judgment_note) > 100:
        return jsonify({'error': 'หมายเหตุ / เงื่อนไขพิเศษเพิ่มเติมต้องไม่เกิน 100 ตัวอักษร'}), 400

    if filing_date and judgment_date and judgment_date <= filing_date:
        return jsonify({'error': 'วันที่พิพากษาต้องมากกว่าวันที่ยื่นฟ้อง'}), 400

    # วันครบกำหนดงวดแรกเป็นวันเดียวกับวันพิพากษาได้ทั้งสองประเภท
    if judgment_date and first_due_date and first_due_date < judgment_date:
        return jsonify({'error': 'วันครบกำหนดงวดแรกต้องไม่น้อยกว่าวันที่พิพากษา'}), 400

    int_rate     = float(data.get('interest_rate', 0) or 0)
    default_rate = float(data.get('default_interest_rate', 0) or 0)
    total_debt = float(data.get('total_debt', 0) or 0)
    principal = float(data.get('principal', 0) or 0)
    court_fee = float(data.get('court_fee', 0) or 0)
    lawyer_fee = float(data.get('lawyer_fee', 0) or 0)

    # ใช้ BE เป็น source of truth สำหรับยอดหนี้ส่วนต่าง
    judgment_difference = calculate_judgment_difference({
        'court_fee': court_fee,
        'lawyer_fee': lawyer_fee,
        'total_debt': total_debt,
        'principal': principal,
    })

    try:
        first_due_d   = date.fromisoformat(first_due_date)
        last_due_date = (first_due_d + relativedelta(months=int(data['installment_count']) - 1)).isoformat()
    except Exception:
        last_due_date = data.get('last_due_date')

    db.execute('''
        UPDATE customers SET
            case_status           = ?,
            red_case_no           = ?,
            judgment_date         = ?,
            judgment_note         = ?,
            total_debt            = ?,
            principal             = ?,
            judgment_difference   = ?,
            interest_rate         = ?,
            court_fee             = ?,
            lawyer_fee            = ?,
            installment_count     = ?,
            default_interest_rate = ?,
            first_due_date        = ?,
            last_due_date         = ?,
            installment_1         = ?,
            installment_2         = ?,
            installment_3         = ?,
            installment_4         = ?,
            updated_at            = CURRENT_TIMESTAMP
        WHERE account_no = ? AND is_deleted = 0
    ''', (
        data['judgment_type'],
        red_case_no,
        judgment_date,
        judgment_note or None,
        total_debt,
        principal,
        judgment_difference,
        int_rate,
        court_fee,
        lawyer_fee,
        int(data['installment_count']),
        default_rate,
        first_due_date,
        last_due_date,
        float(data['installment_1']),
        float(data.get('installment_2', 0)),
        float(data.get('installment_3', 0)),
        float(data.get('installment_4', 0)),
        account_no
    ))

    _log_case_status(db, account_no, cus['case_status'], data['judgment_type'], user['id'], 'บันทึกคำพิพากษา')

    # บันทึก edit history
    JUDGMENT_LABELS = {
        'judgment_type':         'ประเภทคำพิพากษา',
        'red_case_no':           'คดีหมายเลขแดงที่',
        'judgment_date':         'วันพิพากษา',
        'judgment_note':         'หมายเหตุ / เงื่อนไขพิเศษเพิ่มเติม',
        'total_debt':            'ยอดหนี้รวม',
        'principal':              'เงินต้น',
        'judgment_difference':    'ยอดหนี้ส่วนต่าง',
        'interest_rate':          'อัตราดอกเบี้ย',
        'court_fee':             'ค่าธรรมเนียมศาล',
        'lawyer_fee':            'ค่าทนาย',
        'installment_count':     'จำนวนงวด',
        'default_interest_rate': 'ดอกเบี้ยผิดนัด',
        'first_due_date':        'วันครบกำหนดงวดแรก',
        'last_due_date':         'วันครบกำหนดงวดสุดท้าย',
        'installment_1':         'ค่างวด 1',
        'installment_2':         'ค่างวด 2',
        'installment_3':         'ค่างวด 3',
        'installment_4':         'ค่างวด 4',
    }
    data['red_case_no'] = red_case_no
    data['judgment_note'] = judgment_note
    data['judgment_difference'] = judgment_difference
    data['last_due_date'] = last_due_date
    judgment_changes = {}
    for field, label in JUDGMENT_LABELS.items():
        old_val = cus.get(field) if field != 'judgment_type' else cus.get('case_status')
        new_val = data.get(field)
        if str(old_val) != str(new_val):
            judgment_changes[field] = {'label': label, 'from': old_val, 'to': new_val}

    if judgment_changes:
        db.execute(
            'INSERT INTO customer_edits (customer_id, account_no, edited_by, changes) VALUES (?, ?, ?, ?)',
            (cus['id'], account_no, user['id'], json.dumps(judgment_changes, ensure_ascii=False))
        )

    db.commit()

    refresh_customer_status(account_no, db)

    return jsonify({'message': 'บันทึกคำพิพากษาสำเร็จ', 'account_no': account_no, 'case_status': data['judgment_type']}), 200


@bp.route('/<account_no>/enforcement', methods=['PATCH'])
def update_enforcement(account_no):
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    db  = get_db()
    cus = db.execute(
        'SELECT * FROM customers WHERE account_no = ? AND is_deleted = 0', (account_no,)
    ).fetchone()
    if not cus:
        return jsonify({'error': 'ไม่พบข้อมูล'}), 404
    cus = dict(cus)

    if not _can_record_enforcement_role(user):
        return jsonify({'error': 'ไม่มีสิทธิ์แก้ไขข้อมูลในสถานะนี้'}), 403

    data = request.get_json()

    required = ['enforcement_judgment_date']
    missing = [f for f in required if not data.get(f)]
    if missing:
        return jsonify({'error': f'กรุณากรอกข้อมูลให้ครบ: {", ".join(missing)}'}), 400
    if data.get('enforcement_judgment_date') > date.today().isoformat():
        return jsonify({'error': 'วันที่ของหมายบังคับคดีต้องไม่เป็นวันที่ในอนาคต'}), 400
    if not cus.get('red_case_no'):
        return jsonify({'error': 'ไม่พบคดีหมายเลขแดงที่สำหรับบันทึกหมายบังคับคดี'}), 400

    allowed = ['พิพากษาตามยอม', 'พิพากษาฝ่ายเดียว']
    if cus['case_status'] not in allowed:
        return jsonify({'error': f'ไม่สามารถบันทึกหมายบังคับคดีได้ สถานะปัจจุบันคือ {cus["case_status"]}'}), 400

    payments = db.execute(
        'SELECT * FROM payments WHERE account_no = ? ORDER BY payment_date ASC', (account_no,)
    ).fetchall()
    payments = [dict(p) for p in payments]

    snap = get_snapshot_at_date(cus, payments, date.today().isoformat())
    if not snap:
        return jsonify({'error': 'ไม่สามารถคำนวณสถานะได้'}), 400

    if not _can_record_enforcement(cus, snap):
        return jsonify({'error': 'ลูกหนี้รายนี้ยังไม่เข้าเงื่อนไขบันทึกหมายบังคับคดี'}), 400

    db.execute('''
        UPDATE customers SET
            case_status               = 'บังคับคดี',
            enforcement_order_no      = ?,
            enforcement_judgment_date = ?,
            enforcement_received_date = ?,
            enforcement_recorded_by   = ?,
            enforcement_recorded_at   = CURRENT_TIMESTAMP,
            updated_at                = CURRENT_TIMESTAMP
        WHERE account_no = ? AND is_deleted = 0
    ''', (
        cus['red_case_no'],
        data['enforcement_judgment_date'],
        None,
        user['id'],
        account_no
    ))

    _log_case_status(
        db,
        account_no,
        cus['case_status'],
        'บังคับคดี',
        user['id'],
        f'บันทึกหมายบังคับคดี วันที่ของหมาย {data["enforcement_judgment_date"]}'
    )

    enforcement_changes = {}
    enforcement_new_values = {
        'case_status': 'บังคับคดี',
        'enforcement_order_no': cus['red_case_no'],
        'enforcement_judgment_date': data['enforcement_judgment_date'],
        'enforcement_received_date': None,
    }
    enforcement_labels = {
        'case_status': 'สถานะคดี',
        'enforcement_order_no': 'หมายเลขบังคับคดี',
        'enforcement_judgment_date': 'วันที่ของหมายบังคับคดี',
        'enforcement_received_date': 'วันที่ได้รับหมายบังคับคดี',
    }
    for field, new_val in enforcement_new_values.items():
        old_val = cus.get(field)
        if str(old_val) != str(new_val):
            enforcement_changes[field] = {
                'label': enforcement_labels[field],
                'from': old_val,
                'to': new_val,
            }

    if enforcement_changes:
        db.execute(
            'INSERT INTO customer_edits (customer_id, account_no, edited_by, changes) VALUES (?, ?, ?, ?)',
            (cus['id'], account_no, user['id'], json.dumps(enforcement_changes, ensure_ascii=False))
        )

    refresh_customer_list_cache(account_no, db=db, commit=False)
    db.commit()

    return jsonify({'message': 'บันทึกหมายบังคับคดีสำเร็จ', 'account_no': account_no}), 200


@bp.route('/<account_no>/status-logs', methods=['GET'])
def get_status_logs(account_no):
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    db   = get_db()
    rows = db.execute('''
        SELECT csl.*, u.display_name as changed_by_name
        FROM case_status_logs csl
        LEFT JOIN users u ON u.id = csl.changed_by
        WHERE csl.account_no = ?
        ORDER BY csl.changed_at ASC
    ''', (account_no,)).fetchall()

    return jsonify({'logs': [dict(r) for r in rows]}), 200


@bp.route('/bulk-judgment', methods=['POST'])
def bulk_judgment():
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401
    if user['role'] != 'admin':
        return jsonify({'error': 'เฉพาะ Admin เท่านั้น'}), 403

    if 'file' not in request.files:
        return jsonify({'error': 'กรุณาแนบไฟล์ Excel'}), 400

    file = request.files['file']
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': 'รองรับเฉพาะไฟล์ .xlsx'}), 400

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({'error': f'ไม่สามารถอ่านไฟล์ได้: {str(e)}'}), 400

    db      = get_db()
    results = []
    success = skip = error = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if all(v is None for v in row):
            continue

        account_no    = str(row[0]).strip() if row[0] else None
        judgment_type = str(row[1]).strip() if row[1] else None

        if not account_no:
            results.append({'row': row_idx, 'status': 'error', 'message': 'account_no ว่างเปล่า'})
            error += 1
            continue

        if judgment_type not in ['พิพากษาตามยอม', 'พิพากษาฝ่ายเดียว']:
            results.append({'row': row_idx, 'account_no': account_no, 'status': 'error',
                            'message': 'judgment_type ต้องเป็น พิพากษาตามยอม หรือ พิพากษาฝ่ายเดียว'})
            error += 1
            continue

        cus = db.execute(
            'SELECT * FROM customers WHERE account_no = ? AND is_deleted = 0', (account_no,)
        ).fetchone()
        if not cus:
            results.append({'row': row_idx, 'account_no': account_no, 'status': 'error', 'message': 'ไม่พบในระบบ'})
            error += 1
            continue
        cus = dict(cus)

        if cus['case_status'] != 'ยื่นฟ้อง':
            results.append({'row': row_idx, 'account_no': account_no, 'status': 'skip',
                            'message': f'สถานะปัจจุบันคือ {cus["case_status"]} ไม่ใช่ ยื่นฟ้อง'})
            skip += 1
            continue

        try:
            from dateutil.relativedelta import relativedelta as _rd
            red_case_no    = _normalize_red_case_no(row[2] if len(row) > 2 else None)
            if not red_case_no:
                raise ValueError('คดีหมายเลขแดงที่ว่างเปล่าหรือรูปแบบไม่ถูกต้อง เช่น ผบ1234/2567, ผบE814/2569 หรือ พE325/2568')
            judgment_date  = str(row[3]).strip() if len(row) > 3 and row[3] else None
            total_debt     = float(row[4] or 0)
            principal      = float(row[5] or 0)
            interest_rate  = float(row[6] or 0)
            court_fee      = float(row[7] or 0)
            lawyer_fee     = float(row[8] or 0)
            judgment_difference = calculate_judgment_difference({
                'court_fee': court_fee,
                'lawyer_fee': lawyer_fee,
                'total_debt': total_debt,
                'principal': principal,
            })
            inst_count     = int(float(row[9] or 0))
            first_due_date = str(row[10]).strip() if len(row) > 10 and row[10] else None
            inst1          = float(row[11] or 0)
            inst2          = float(row[12] or 0)
            inst3          = float(row[13] or 0)
            inst4          = float(row[14] or 0)
            default_rate   = float(row[15] or 0)
            judgment_note  = str(row[16]).strip() if len(row) > 16 and row[16] not in (None, '') else ''
            if len(judgment_note) > 100:
                raise ValueError('หมายเหตุ / เงื่อนไขพิเศษเพิ่มเติมต้องไม่เกิน 100 ตัวอักษร')
            if not _parse_iso_date(judgment_date):
                raise ValueError('วันที่พิพากษาต้องเป็นรูปแบบ YYYY-MM-DD เช่น 2026-02-18')
            if not _parse_iso_date(first_due_date):
                raise ValueError('วันครบกำหนดงวดแรกต้องเป็นรูปแบบ YYYY-MM-DD เช่น 2026-03-18')
            _validate_judgment_timeline(
                cus.get('filing_date'),
                judgment_type,
                judgment_date,
                first_due_date,
            )

            last_due_date = (
                date.fromisoformat(first_due_date) + _rd(months=inst_count - 1)
            ).isoformat() if first_due_date else None

            db.execute('''
                UPDATE customers SET
                    case_status           = ?,
                    red_case_no           = ?,
                    judgment_date         = ?,
                    judgment_note         = ?,
                    total_debt            = ?,
                    principal             = ?,
                    judgment_difference   = ?,
                    interest_rate         = ?,
                    court_fee             = ?,
                    lawyer_fee            = ?,
                    installment_count     = ?,
                    default_interest_rate = ?,
                    first_due_date        = ?,
                    last_due_date         = ?,
                    installment_1         = ?,
                    installment_2         = ?,
                    installment_3         = ?,
                    installment_4         = ?,
                    updated_at            = CURRENT_TIMESTAMP
                WHERE account_no = ? AND is_deleted = 0
            ''', (
                judgment_type, red_case_no, judgment_date, judgment_note or None, total_debt, principal,
                judgment_difference,
                interest_rate, court_fee, lawyer_fee, inst_count, default_rate,
                first_due_date, last_due_date, inst1, inst2, inst3, inst4,
                account_no
            ))

            _log_case_status(db, account_no, cus['case_status'], judgment_type, user['id'], 'Bulk import คำพิพากษา')
            results.append({'row': row_idx, 'account_no': account_no, 'status': 'success',
                            'message': f'บันทึกสำเร็จ → {judgment_type}'})
            success += 1

        except Exception as e:
            results.append({'row': row_idx, 'account_no': account_no, 'status': 'error', 'message': str(e)})
            error += 1

    db.commit()

    for r in results:
        if r['status'] == 'success':
            refresh_customer_status(r['account_no'], db)

    return jsonify({
        'summary': {'success': success, 'skip': skip, 'error': error, 'total': success + skip + error},
        'results': results,
    }), 200


@bp.route('/<account_no>/edits', methods=['GET'])
def get_customer_edits(account_no):
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    import json
    db   = get_db()
    rows = db.execute(
        """SELECT ce.*, u.display_name as edited_by_name
           FROM customer_edits ce
           LEFT JOIN users u ON u.id = ce.edited_by
           WHERE ce.account_no = ?
           ORDER BY ce.edited_at DESC""",
        (account_no,)
    ).fetchall()

    edits = []
    for r in rows:
        r = dict(r)
        try:
            r['changes'] = json.loads(r['changes'])
        except Exception:
            r['changes'] = {}
        edits.append(r)

    return jsonify({'edits': edits}), 200


@bp.route('/<account_no>', methods=['DELETE'])
def delete_customer(account_no):
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401
    if user['role'] != 'admin':
        return jsonify({'error': 'Forbidden'}), 403

    db = get_db()
    db.execute(
        'UPDATE customers SET is_deleted = 1, deleted_at = CURRENT_TIMESTAMP WHERE account_no = ?',
        (account_no,)
    )
    db.commit()

    return jsonify({'message': 'Deleted'}), 200


@bp.route('/<account_no>', methods=['PUT'])
def update_customer(account_no):
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    db = get_db()

    existing = db.execute(
        'SELECT id FROM customers WHERE account_no = ? AND is_deleted = 0', (account_no,)
    ).fetchone()
    if not existing:
        return jsonify({'error': 'ไม่พบข้อมูล'}), 404

    old = dict(db.execute(
        'SELECT * FROM customers WHERE account_no = ? AND is_deleted = 0', (account_no,)
    ).fetchone())

    if not _can_edit_customer_detail(user, old.get('case_status')):
        return jsonify({'error': 'ไม่มีสิทธิ์แก้ไขข้อมูลในสถานะนี้'}), 403

    data = request.get_json()

    required = ['filing_date', 'filing_capital', 'judgment_date', 'total_debt', 'principal',
            'interest_rate', 'installment_count', 'first_due_date', 'installment_1']
    if not (old.get('red_case_no') or '').strip():
        required.append('red_case_no')
    missing = [f for f in required if data.get(f) is None]
    if missing:
        return jsonify({'error': f'กรุณากรอกข้อมูลให้ครบ: {", ".join(missing)}'}), 400

    EDITABLE_FIELDS = {
        'filing_date':           'วันที่ยื่นฟ้อง',
        'filing_capital':        'ทุนทรัพย์ที่ฟ้อง',
        'red_case_no':           'คดีหมายเลขแดงที่',
        'judgment_date':         'วันพิพากษา',
        'judgment_note':         'หมายเหตุ / เงื่อนไขพิเศษเพิ่มเติม',
        'total_debt':            'ยอดหนี้รวม',
        'principal':             'เงินต้น',
        'court_fee':             'ค่าธรรมเนียมศาล',
        'lawyer_fee':            'ค่าทนาย',
        'interest_rate':         'อัตราดอกเบี้ย',
        'default_interest_rate': 'ดอกเบี้ยผิดนัด',
        'installment_count':     'จำนวนงวด',
        'first_due_date':        'วันครบกำหนดงวดแรก',
        'last_due_date':         'วันครบกำหนดงวดสุดท้าย',
        'installment_1':         'ค่างวด 1',
        'installment_2':         'ค่างวด 2',
        'installment_3':         'ค่างวด 3',
        'installment_4':         'ค่างวด 4',
        'judgment_difference': 'ยอดหนี้ส่วนต่าง',
    }

    filing_capital_text = str(data.get('filing_capital') or '').replace(',', '').strip()
    if not re.fullmatch(r'\d+(\.\d{1,2})?', filing_capital_text):
        return jsonify({'error': 'ทุนทรัพย์ที่ฟ้องต้องเป็นตัวเลข และมีทศนิยมได้สูงสุด 2 ตำแหน่ง'}), 400

    filing_capital = round(float(filing_capital_text), 2)
    if filing_capital <= 0:
        return jsonify({'error': 'ทุนทรัพย์ที่ฟ้องต้องมากกว่า 0'}), 400

    old_red_case_no = str(old.get('red_case_no') or '').strip()
    normalized_old_red_case_no = _normalize_red_case_no(old_red_case_no)
    incoming_red_case_no = _normalize_red_case_no(data.get('red_case_no'))
    red_case_no = incoming_red_case_no
    if old_red_case_no:
        if incoming_red_case_no and normalized_old_red_case_no and incoming_red_case_no != normalized_old_red_case_no:
            return jsonify({'error': 'คดีหมายเลขแดงที่ถูกบันทึกครั้งแรกแล้ว ไม่อนุญาตให้แก้ไข'}), 400
        red_case_no = old_red_case_no

    judgment_note = str(data.get('judgment_note') or '').strip()
    if not red_case_no:
        return jsonify({'error': 'คดีหมายเลขแดงที่ต้องอยู่ในรูปแบบ: ตัวย่อประเภทคดีติดเลขที่/ปี พ.ศ. เช่น ผบ1234/2567, ผบE814/2569 หรือ พE325/2568'}), 400
    if len(judgment_note) > 100:
        return jsonify({'error': 'หมายเหตุ / เงื่อนไขพิเศษเพิ่มเติมต้องไม่เกิน 100 ตัวอักษร'}), 400

    new_vals = {
        'filing_date':           data['filing_date'],
        'filing_capital':        filing_capital,
        'red_case_no':           red_case_no,
        'judgment_date':         data['judgment_date'],
        'judgment_note':         judgment_note,
        'total_debt':            float(data['total_debt']),
        'principal':             float(data['principal']),
        'court_fee':             float(data.get('court_fee', 0)),
        'lawyer_fee':            float(data.get('lawyer_fee', 0)),
        'interest_rate':         float(data['interest_rate']),
        'default_interest_rate': float(data.get('default_interest_rate', 0)),
        'installment_count':     int(data['installment_count']),
        'first_due_date':        data['first_due_date'],
        'last_due_date':         data.get('last_due_date'),
        'installment_1':         float(data['installment_1']),
        'installment_2':         float(data.get('installment_2', 0)),
        'installment_3':         float(data.get('installment_3', 0)),
        'installment_4':         float(data.get('installment_4', 0)),
    }
    new_vals['judgment_difference'] = calculate_judgment_difference(new_vals)

    import json
    changes = {}
    for field, label in EDITABLE_FIELDS.items():
        old_val = old.get(field)
        new_val = new_vals.get(field)
        if str(old_val) != str(new_val):
            changes[field] = {'label': label, 'from': old_val, 'to': new_val}

    try:
        db.execute('ALTER TABLE customers ADD COLUMN updated_at DATETIME')
    except Exception:
        pass

    db.execute('''
        UPDATE customers SET
            filing_date           = ?,
            filing_capital        = ?,
            red_case_no           = ?,
            judgment_date         = ?,
            judgment_note         = ?,
            total_debt            = ?,
            principal             = ?,
            judgment_difference   = ?,
            interest_rate         = ?,
            court_fee             = ?,
            lawyer_fee            = ?,
            installment_count     = ?,
            default_interest_rate = ?,
            first_due_date        = ?,
            last_due_date         = ?,
            installment_1         = ?,
            installment_2         = ?,
            installment_3         = ?,
            installment_4         = ?,
            updated_at            = CURRENT_TIMESTAMP
        WHERE account_no = ? AND is_deleted = 0
    ''', (
        new_vals['filing_date'],
        new_vals['filing_capital'],
        new_vals['red_case_no'],
        new_vals['judgment_date'],
        new_vals['judgment_note'] or None,
        new_vals['total_debt'],
        new_vals['principal'],
        new_vals['judgment_difference'],
        new_vals['interest_rate'],
        new_vals['court_fee'],
        new_vals['lawyer_fee'],
        new_vals['installment_count'],
        new_vals['default_interest_rate'],
        new_vals['first_due_date'],
        new_vals['last_due_date'],
        new_vals['installment_1'],
        new_vals['installment_2'],
        new_vals['installment_3'],
        new_vals['installment_4'],
        account_no
    ))

    if changes:
        db.execute(
            'INSERT INTO customer_edits (customer_id, account_no, edited_by, changes) VALUES (?, ?, ?, ?)',
            (old['id'], account_no, user['id'], json.dumps(changes, ensure_ascii=False))
        )

    db.commit()

    refresh_customer_status(account_no, db)

    return jsonify({'message': 'อัพเดทข้อมูลสำเร็จ', 'account_no': account_no, 'changed_fields': len(changes)}), 200
