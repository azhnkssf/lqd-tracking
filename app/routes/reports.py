import io
import openpyxl
from decimal import Decimal, ROUND_DOWN, InvalidOperation
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.worksheet import Worksheet
from flask import Blueprint, current_app, request, jsonify, send_file
from app.database import get_db
from app.services.auth_service import get_user_by_token
from app.services.report_service import (
    REPORT_MODE_CORRECTED,
    REPORT_MODE_NORMAL,
    get_snapshot_at_date,
    get_report30_snapshot_at_date,
    _get_report_group,
    _build_report30_row_from_db,
    _build_report31_row,
    _build_report11_row,
    _is_customer_effective_after_report,
    _attach_case_status_log_context,
    _future_effective_reason,
    _build_not_generated_row,
    _build_customer_as_of_report_date,
    apply_correction_warning_remark,
    build_correction_summary,
    build_retroactive_alerts,
    REPORT30_HEADERS,
    build_report30_export_values,
)

bp = Blueprint('reports', __name__, url_prefix='/api/report')


def get_current_user():
    token = request.cookies.get(current_app.config.get('AUTH_COOKIE_NAME', 'token')) or request.headers.get('Authorization', '').replace('Bearer ', '')
    return get_user_by_token(token)


@bp.before_request
def block_superadmin_from_report_api():
    user = get_current_user()
    if user and user['role'] == 'superadmin':
        return jsonify({'error': 'Superadmin is limited to user management.'}), 403


def trunc_money(v, default=0):
    if v is None or v == '':
        return default

    try:
        return int(Decimal(str(v)).to_integral_value(rounding=ROUND_DOWN))
    except (InvalidOperation, ValueError, TypeError):
        return default


def build_reconcile_summary(db_total, report_30, report_31, report_11, alerts, missing_db, not_generated, retroactive_alerts=None):
    generated_total = len(report_30) + len(report_31) + len(report_11)
    not_generated_total = len(not_generated)
    missing_total = len(missing_db)
    retroactive_total = len(retroactive_alerts or [])
    report_30_status_counts = {}

    for row in report_30:
        case_status = row.get('case_status') or row.get('status_label') or '-'
        report_30_status_counts[case_status] = report_30_status_counts.get(case_status, 0) + 1

    return {
        'report_30'          : len(report_30),
        'report_30_status_counts': report_30_status_counts,
        'report_31'          : len(report_31),
        'report_11'          : len(report_11),
        'alerts'             : len(alerts),
        'retroactive_alerts' : retroactive_total,
        'missing_db'         : missing_total,
        'db_total'           : db_total,
        'generated_total'    : generated_total,
        'not_generated'      : not_generated_total,
        'reconcile_total'    : generated_total + not_generated_total + missing_total,
        'reconcile_matched'  : db_total == generated_total + not_generated_total + missing_total,
    }


@bp.route('/export/<report_type>', methods=['POST'])
def export_report(report_type):
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    data        = request.get_json()
    rows        = data.get('rows', [])
    report_date = data.get('report_date', '')

    def fmt_date_excel(val, default=''):
        if not val:
            return default
        s = str(val).strip().split('T')[0].split(' ')[0]
        if not s or s.lower() in ['none', 'null', '-']:
            return default

        if len(s) == 8 and s.isdigit():
            return s

        parts = s.split('-')
        if len(parts) == 3 and len(parts[0]) == 4:
            return f"{parts[0]}{parts[1]}{parts[2]}"

        parts = s.split('/')
        if len(parts) == 3:
            if len(parts[0]) == 4:
                return f"{parts[0]}{parts[1].zfill(2)}{parts[2].zfill(2)}"
            return f"{parts[2]}{parts[1].zfill(2)}{parts[0].zfill(2)}"

        return s

    def fmt_acc(val):
        s = str(val).strip() if val else ''
        digits = ''.join(ch for ch in s if ch.isdigit())
        return digits or '-'

    wb = openpyxl.Workbook()
    ws: Worksheet = wb.active

    header_fill = PatternFill(start_color='1e3a8a', end_color='1e3a8a', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    center      = Alignment(horizontal='center')

    fill_30 = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    fill_31 = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')

    if report_type == '30':
        ws.title = 'Report Status 30'
        headers = REPORT30_HEADERS
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        def fmt_num_30(v, decimals=2):
            return trunc_money(v, default=0)

        for r in rows:
            ws.append(build_report30_export_values(r, fmt_acc, fmt_date_excel, fmt_num_30))
            for cell in ws[ws.max_row]:
                cell.fill = fill_30

        for row in ws.iter_rows(min_row=2):
            row[2].number_format = '#,##0'  # ยอดเงินส่งฟ้อง
            row[3].number_format = '#,##0'  # ยอดเงินที่ศาลตัดสิน
            row[7].number_format = '#,##0'  # ค่างวดคงเหลือ

    elif report_type == '31':
        ws.title = 'Report Status 31'
        headers = [
            'เลขที่บัญชี', 'Amount Owed', 'Amount Past Due',
            'DPD (เดือน)', 'Default Date', 'Installment Amount',
            'จำนวนงวด', 'Frequency', 'Maturity Date', 'Remark'
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        def fmt_num(v, decimals=2):
            return trunc_money(v, default=0)

        for r in rows:
            ws.append([
                fmt_acc(r.get('account_no')),
                fmt_num(r.get('amount_owed'), 2),
                fmt_num(r.get('amount_past_due'), 2),
                int(r.get('dpd') or 0),
                fmt_date_excel(r.get('default_date'), default='19000101'),
                fmt_num(r.get('installment_amount'), 2),
                int(r.get('installment_count') or 0),
                str(r.get('frequency') or '00'),
                fmt_date_excel(r.get('maturity_date')),
                r.get('remark') or '',
            ])
            for cell in ws[ws.max_row]:
                cell.fill = fill_31
                
        for row in ws.iter_rows(min_row=2):
            row[1].number_format = '#,##0'
            row[2].number_format = '#,##0'
            row[5].number_format = '#,##0'

    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'Report_Status_{report_type}_{report_date}.xlsx'
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@bp.route('/export-all', methods=['POST'])
def export_all_reports():
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    data = request.get_json() or {}

    report_date = data.get('report_date', '')
    report_30   = data.get('report_30', []) or []
    report_31   = data.get('report_31', []) or []
    report_11   = data.get('report_11', []) or []
    alerts      = data.get('alerts', []) or []
    missing_db  = data.get('missing_db', []) or []
    not_generated = data.get('not_generated', []) or []
    report_mode = data.get('report_mode') or REPORT_MODE_NORMAL

    if not any([report_30, report_31, report_11, alerts, missing_db, not_generated]):
        return jsonify({'error': 'ไม่มีข้อมูลสำหรับ Export'}), 400

    def fmt_date_excel(val, default=''):
        if not val:
            return default
        s = str(val).strip().split('T')[0].split(' ')[0]
        if not s or s.lower() in ['none', 'null', '-']:
            return default

        if len(s) == 8 and s.isdigit():
            return s

        parts = s.split('-')
        if len(parts) == 3 and len(parts[0]) == 4:
            return f"{parts[0]}{parts[1]}{parts[2]}"

        parts = s.split('/')
        if len(parts) == 3:
            if len(parts[0]) == 4:
                return f"{parts[0]}{parts[1].zfill(2)}{parts[2].zfill(2)}"
            return f"{parts[2]}{parts[1].zfill(2)}{parts[0].zfill(2)}"

        return s

    def fmt_num(v, decimals=2):
        return trunc_money(v, default=0)

    def fmt_acc(val):
        s = str(val).strip() if val else ''
        digits = ''.join(ch for ch in s if ch.isdigit())
        return digits or '-'

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color='1E1B4B', end_color='1E1B4B', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    center      = Alignment(horizontal='center', vertical='center')
    left        = Alignment(horizontal='left', vertical='center')

    fill_30      = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    fill_31      = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
    fill_33      = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
    fill_alert   = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
    fill_missing = PatternFill(start_color='FFEDD5', end_color='FFEDD5', fill_type='solid')
    fill_skipped = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')

    def style_header(ws):
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

    def autofit(ws):
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

    def paint_row(ws, fill):
        for cell in ws[ws.max_row]:
            cell.fill = fill
            cell.alignment = left

    if report_30:
        ws = wb.create_sheet('Report 30')
        ws.append(REPORT30_HEADERS)
        style_header(ws)

        for r in report_30:
            ws.append(build_report30_export_values(r, fmt_acc, fmt_date_excel, fmt_num))
            paint_row(ws, fill_30)

        for row in ws.iter_rows(min_row=2):
            row[2].number_format = '#,##0'  # ยอดเงินส่งฟ้อง
            row[3].number_format = '#,##0'  # ยอดเงินที่ศาลตัดสิน
            row[7].number_format = '#,##0'  # ค่างวดคงเหลือ

        autofit(ws)

    if report_31:
        ws = wb.create_sheet('Report 31')
        ws.append([
            'เลขที่บัญชี',
            'Amount Owed',
            'Amount Past Due',
            'DPD (เดือน)',
            'Default Date',
            'Installment Amount',
            'จำนวนงวด',
            'Frequency',
            'Maturity Date',
            'Remark',
        ])
        style_header(ws)

        for r in report_31:
            ws.append([
                fmt_acc(r.get('account_no')),
                fmt_num(r.get('amount_owed'), 2),
                fmt_num(r.get('amount_past_due'), 2),
                int(r.get('dpd') or 0),
                fmt_date_excel(r.get('default_date'), default='19000101'),
                fmt_num(r.get('installment_amount'), 2),
                int(r.get('installment_count') or 0),
                str(r.get('frequency') or '00'),
                fmt_date_excel(r.get('maturity_date')),
                r.get('remark') or '',
            ])
            paint_row(ws, fill_31)

        for row in ws.iter_rows(min_row=2):
            row[1].number_format = '#,##0'
            row[2].number_format = '#,##0'
            row[5].number_format = '#,##0'

        autofit(ws)

    if report_11:
        ws = wb.create_sheet('Status 11')
        ws.append([
            'เลขที่บัญชี',
            'วันที่ปิดบัญชี',
            'NCB Status',
        ])
        style_header(ws)

        for r in report_11:
            ws.append([
                fmt_acc(r.get('account_no')),
                fmt_date_excel(r.get('closed_date')),
                r.get('ncb_status') or '11',
            ])
            paint_row(ws, fill_33)

        autofit(ws)

    if alerts:
        ws = wb.create_sheet('Alerts')
        ws.append([
            'เลขที่บัญชี',
            'สถานะ',
            'วันที่ยื่นฟ้อง',
            'หมายเหตุ',
        ])
        style_header(ws)

        for r in alerts:
            ws.append([
                fmt_acc(r.get('account_no')),
                r.get('status') or '-',
                fmt_date_excel(r.get('filing_date')),
                r.get('message') or '-',
            ])
            paint_row(ws, fill_alert)

        autofit(ws)

    if missing_db:
        ws = wb.create_sheet('Missing DB')
        ws.append([
            'เลขที่บัญชี',
            'ชื่อลูกค้า',
            'หมายเหตุ',
        ])
        style_header(ws)

        for r in missing_db:
            ws.append([
                fmt_acc(r.get('account_no')),
                r.get('name') or '-',
                r.get('message') or '-',
            ])
            paint_row(ws, fill_missing)

        autofit(ws)

    if not_generated:
        ws = wb.create_sheet('Not Generated')
        ws.append([
            'เลขที่บัญชี',
            'ชื่อลูกค้า',
            'สถานะปัจจุบัน',
            'วันที่ยื่นฟ้อง',
            'วันที่พิพากษา',
            'วันที่บังคับคดี',
            'วันที่ปิดบัญชี',
            'วันที่มีผล',
            'Report Date',
            'Reason Code',
            'เหตุผล',
        ])
        style_header(ws)

        for r in not_generated:
            ws.append([
                fmt_acc(r.get('account_no')),
                r.get('name') or '-',
                r.get('case_status') or '-',
                fmt_date_excel(r.get('filing_date')),
                fmt_date_excel(r.get('judgment_date')),
                fmt_date_excel(r.get('enforcement_date')),
                fmt_date_excel(r.get('closed_date')),
                fmt_date_excel(r.get('effective_date')),
                fmt_date_excel(r.get('report_date') or report_date),
                r.get('reason_code') or '-',
                r.get('reason') or '-',
            ])
            paint_row(ws, fill_skipped)

        autofit(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    suffix = '_Corrected' if report_mode == REPORT_MODE_CORRECTED else ''
    filename = f'Report_All_{report_date}{suffix}.xlsx'
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@bp.route('/generate-db', methods=['POST'])
def generate_report_db():
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    return _generate_report_db_from_data(request.get_json() or {}, user)


def _generate_report_db_from_data(data, user):
    report_date  = data.get('report_date')
    report_mode  = data.get('report_mode') or REPORT_MODE_NORMAL
    corrected_scope = data.get('corrected_scope') or 'pending_only'
    status_types = ['30', '31']

    if not report_date:
        return jsonify({'error': 'กรุณาระบุวันที่ขอ Report'}), 400
    if report_mode not in (REPORT_MODE_NORMAL, REPORT_MODE_CORRECTED):
        return jsonify({'error': 'report_mode ไม่ถูกต้อง'}), 400
    if corrected_scope != 'pending_only':
        return jsonify({'error': 'corrected_scope ไม่ถูกต้อง'}), 400

    db = get_db()

    customers = db.execute(
        'SELECT * FROM customers WHERE is_deleted = 0'
    ).fetchall()

    report_30 = []
    report_31 = []
    report_11 = []
    not_generated = []
    retroactive_alerts = []
    db_total = len(customers)

    for cus_row in customers:
        cus         = dict(cus_row)
        cus         = _attach_case_status_log_context(cus, db)
        account_no  = cus['account_no']

        # Keep latest DB customer only for retroactive alert diagnostics.
        latest_cus = cus

        customer_alerts = build_retroactive_alerts(
            latest_cus,
            db=db,
            report_date_str=report_date,
            include_marked=True,
        )
        retroactive_alerts.extend(customer_alerts)
        report_month = str(report_date or '')[:7]
        customer_alerts_for_report = [
            alert for alert in customer_alerts
            if str(alert.get('affected_report_month') or '') == report_month
        ]
        pending_customer_alerts_for_report = [
            alert for alert in customer_alerts_for_report
            if not alert.get('marked')
        ]

        if (
            report_mode == REPORT_MODE_CORRECTED
            and corrected_scope == 'pending_only'
            and not pending_customer_alerts_for_report
        ):
            continue

        # Use customer status as of report_date for report classification.
        # Example: current status is บังคับคดี, but enforcement date is after report_date.
        # Then rollback to previous status, e.g. พิพากษาฝ่ายเดียว / พิพากษาตามยอม.
        alerts_for_snapshot = (
            pending_customer_alerts_for_report
            if report_mode == REPORT_MODE_CORRECTED
            else customer_alerts
        )
        cus = _build_customer_as_of_report_date(
            cus,
            report_date,
            report_mode=report_mode,
            retroactive_alerts=alerts_for_snapshot,
        )
        case_status = (cus.get('case_status') or 'ยื่นฟ้อง').strip()

        payments = None
        if case_status == 'ปิดบัญชี':
            payments = db.execute(
                'SELECT * FROM payments WHERE account_no = ? ORDER BY payment_date ASC',
                (account_no,)
            ).fetchall()
            payments = [dict(p) for p in payments]

        if _is_customer_effective_after_report(cus, report_date, payments):
            reason = _future_effective_reason(cus, report_date, payments)
            if reason:
                not_generated.append(_build_not_generated_row(
                    cus, reason[0], reason[1], report_date, reason[2], payments
                ))
            continue

        if case_status == 'ปิดบัญชี':
            report_11.append(_build_report11_row(account_no, cus, payments))
            continue

        if case_status == 'ยื่นฟ้อง':
            if '30' in status_types:
                row = _build_report30_row_from_db(
                    account_no,
                    case_status,
                    cus.get('filing_date'),
                    cus.get('filing_capital'),
                    cus,
                    None,
                    report_date,
                    payments=[]
                )
                if report_mode == REPORT_MODE_NORMAL:
                    row = apply_correction_warning_remark(row, cus)
                report_30.append(row)
            else:
                not_generated.append(_build_not_generated_row(
                    cus,
                    'REPORT_TYPE_NOT_SELECTED',
                    'Report 30 ไม่ได้ถูกเลือกตอน Generate',
                    report_date,
                ))
            continue

        if payments is None:
            payments = db.execute(
                'SELECT * FROM payments WHERE account_no = ? ORDER BY payment_date ASC',
                (account_no,)
            ).fetchall()
            payments = [dict(p) for p in payments]

        try:
            snap = get_snapshot_at_date(cus, payments, report_date)
        except Exception:
            snap = None

        if case_status == 'พิพากษาฝ่ายเดียว':
            if '30' in status_types:
                row = _build_report30_row_from_db(
                    account_no,
                    case_status,
                    cus.get('filing_date'),
                    cus.get('filing_capital'),
                    cus,
                    snap,
                    report_date,
                    payments=payments
                )
                if report_mode == REPORT_MODE_NORMAL:
                    row = apply_correction_warning_remark(row, cus)
                report_30.append(row)
            else:
                not_generated.append(_build_not_generated_row(
                    cus,
                    'REPORT_TYPE_NOT_SELECTED',
                    'Report 30 ไม่ได้ถูกเลือกตอน Generate',
                    report_date,
                ))
            continue

        if case_status == 'บังคับคดี':
            if '30' in status_types:
                report30_snap = get_report30_snapshot_at_date(cus, payments, report_date)
                row = _build_report30_row_from_db(
                    account_no,
                    case_status,
                    cus.get('filing_date'),
                    cus.get('filing_capital'),
                    cus,
                    report30_snap,
                    report_date,
                    payments=payments
                )
                if report_mode == REPORT_MODE_NORMAL:
                    row = apply_correction_warning_remark(row, cus)
                report_30.append(row)
            else:
                not_generated.append(_build_not_generated_row(
                    cus,
                    'REPORT_TYPE_NOT_SELECTED',
                    'Report 30 ไม่ได้ถูกเลือกตอน Generate',
                    report_date,
                ))
            continue

        if not snap:
            not_generated.append(_build_not_generated_row(
                cus,
                'NO_SNAPSHOT',
                'ไม่สามารถคำนวณ snapshot ณ วันที่ Report ได้',
                report_date,
            ))
            continue

        group = _get_report_group(cus, snap)

        if group == '11':
            report_11.append(_build_report11_row(account_no, cus, payments))

        elif group == '30' and '30' in status_types:
            row = _build_report30_row_from_db(
                account_no,
                case_status,
                cus.get('filing_date'),
                cus.get('filing_capital'),
                cus,
                snap,
                report_date,
                payments=payments
            )
            if report_mode == REPORT_MODE_NORMAL:
                row = apply_correction_warning_remark(row, cus)
            report_30.append(row)

        elif group == '31' and '31' in status_types:
            row = _build_report31_row(account_no, cus, snap, report_date)
            if report_mode == REPORT_MODE_NORMAL:
                row = apply_correction_warning_remark(row, cus)
            report_31.append(row)
        elif group in ['30', '31']:
            not_generated.append(_build_not_generated_row(
                cus,
                'REPORT_TYPE_NOT_SELECTED',
                f'Report {group} ไม่ได้ถูกเลือกตอน Generate',
                report_date,
            ))

    report_month = str(report_date or '')[:7]
    retroactive_alerts_for_report = [
        alert for alert in retroactive_alerts
        if str(alert.get('affected_report_month') or '') == report_month
    ]
    correction_summary = build_correction_summary(retroactive_alerts_for_report, report_date)

    summary = build_reconcile_summary(
        db_total,
        report_30,
        report_31,
        report_11,
        [],
        [],
        not_generated,
        retroactive_alerts_for_report,
    )
    summary['correction_summary'] = correction_summary

    db.execute(
        """INSERT INTO report_logs
           (generated_by, report_date, filename, status_types,
            count_30, count_31, count_33, count_alerts, count_missing,
            count_skipped, count_db_total, count_generated_total)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
        (user['id'], report_date, '[DB Corrected Report]' if report_mode == REPORT_MODE_CORRECTED else '[DB Report]', ','.join(status_types),
         len(report_30), len(report_31), len(report_11), len(retroactive_alerts_for_report), 0,
         len(not_generated), db_total, summary['generated_total'])
    )
    db.commit()

    return jsonify({
        'report_date' : report_date,
        'report_mode' : report_mode,
        'summary'     : summary,
        'correction_summary': correction_summary,
        'report_30'   : report_30,
        'report_31'   : report_31,
        'report_11'   : report_11,
        'alerts'      : [],
        'retroactive_alerts': retroactive_alerts_for_report,
        'missing_db'  : [],
        'not_generated': not_generated,
    }), 200


@bp.route('/generate-corrected', methods=['POST'])
def generate_corrected_report_db():
    data = request.get_json() or {}
    data['report_mode'] = REPORT_MODE_CORRECTED
    data.setdefault('corrected_scope', 'pending_only')

    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    return _generate_report_db_from_data(data, user)

@bp.route('/history', methods=['GET'])
def get_report_history():
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    db   = get_db()
    rows = db.execute(
        """SELECT rl.*, u.display_name as generated_by_name
           FROM report_logs rl
           LEFT JOIN users u ON u.id = rl.generated_by
           ORDER BY rl.generated_at DESC
           LIMIT 50"""
    ).fetchall()

    return jsonify({'history': [dict(r) for r in rows]}), 200
